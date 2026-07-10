import os
import time
import datetime
import pytz
import pandas as pd
import requests
import smtplib
from functools import wraps
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
import io
import re
from collections import defaultdict
from flask import send_file
import threading
import gc
from dotenv import load_dotenv
import secrets
from urllib.parse import urlparse, urljoin
from werkzeug.security import check_password_hash, generate_password_hash
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.middleware.proxy_fix import ProxyFix
from flask_talisman import Talisman
import logging
import psycopg2
import psycopg2.extras

# Input validation & sanitization
from validators import (
    sanitize_text, validate_mobile, validate_date_str, validate_claim_id,
    validate_status, validate_file_upload, validate_search_type,
    validate_claim_ids_list, validate_reset_token, validate_password_strength,
    validate_username, validate_integer_range, validate_boolean, validate_time_slot,
    MAX_TEXT_SHORT, MAX_TEXT_DEFAULT, MAX_TEXT_LONG,
)

load_dotenv()

# Performance monitoring
from perf_utils import timed_excel_read

# PostgreSQL sync service
try:
    from services.pg_sync import fetch_claims_from_postgres, upsert_claim_to_postgres, test_connection as pg_test_connection
    PG_AVAILABLE = bool(os.environ.get('DATABASE_URL'))
except ImportError:
    PG_AVAILABLE = False
    logging.warning("[PG_SYNC] psycopg2 not installed — PostgreSQL fallback disabled.")

# ----------------------
# CONFIG
# ----------------------
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")

# Email Config
TARGET_EMAIL = "srteam@onsite.co.in"
CC_EMAILS = ["shine.at@onsite.co.in", "akhilmp@myg.in","sachin.kadam@onsite.co.in","shanmugaraja.a@onsite.co.in","akhil.chandran@onsite.co.in","shyla.mariadhasan@onsite.co.in","jasil@myg.in"]
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SENDER_EMAIL = "sarath.k@myg.in"
SENDER_PASSWORD = os.environ.get('SENDER_PASSWORD')

# WhatsApp Notification Cutoff
# Claims registered BEFORE this date will NOT receive WhatsApp notifications.
# Format: YYYY-MM-DD. Change this date as needed.
WHATSAPP_CUTOFF_DATE = datetime.datetime(2026, 7, 9)

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///site.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Rate limiting
limiter = Limiter(get_remote_address, app=app, default_limits=[])

app.secret_key = os.environ.get('FLASK_SECRET_KEY')  # Required for session
app.permanent_session_lifetime = datetime.timedelta(hours=24) # 24 hour session expiry

# Secure session cookie flags
app.config['SESSION_COOKIE_HTTPONLY'] = True   # Prevent JS access to session cookie
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # CSRF mitigation for cross-site requests
app.config['SESSION_COOKIE_SECURE'] = False     # Only send cookie over HTTPS

# Setup Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# Secure Deployment Middleware
# Trust headers from Render Load Balancer
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)
# Enforce HTTPS and add security headers
Talisman(app, content_security_policy=None, force_https=False)

db = SQLAlchemy(app)

# ----------------------
# SECURITY AND THREAT LOGGING
# ----------------------
@app.before_request
def restrict_sensitive_files():
    # Prevent traversal or direct access to sensitive filetypes or directories
    forbidden_strings = ['.env', '.db', '.git', '.yaml', '/instance/']
    
    # Check if requested path contains forbidden substrings
    path = request.path.lower()
    for forbidden in forbidden_strings:
        if forbidden in path:
            logging.warning(f"[THREAT_DETECTED] Blocked path traversal attempt from IP {get_remote_address()} targeting {path}")
            return jsonify({"error": "Forbidden"}), 403

@app.errorhandler(429)
def ratelimit_handler(e):
    logging.warning(f"[THREAT_DETECTED] Rate limit exceeded from IP {get_remote_address()} on route {request.path}: {e.description}")
    return render_template('login.html', error=f"Rate limit exceeded: {e.description}. Try again later."), 429

@app.errorhandler(Exception)
def handle_exception(e):
    logging.error(f"[SYSTEM_ERROR] Exception on {request.path} from IP {get_remote_address()}: {str(e)}", exc_info=True)
    return "Internal Server Error", 500

# ----------------------
# AUTHENTICATION
# ----------------------
USERS = {
    "admin": {
        "password_hash": os.environ.get('ADMIN_PASSWORD_HASH'), 
        "role": "admin", 
        "display": "Admin Manager",
        "email": os.environ.get('ADMIN_EMAIL', 'jasil@myg.in')
    },
    "customercare": {
        "password_hash": os.environ.get('CUSTOMERCARE_PASSWORD_HASH'), 
        "role": "customercare", 
        "display": "Customer Care",
        "email": os.environ.get('CUSTOMERCARE_EMAIL', 'jasil@myg.in')
    }
}

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_logged_in' not in session:
            if request.is_json or request.path.startswith('/api') or request.path == '/lookup-customer':
                return jsonify({'success': False, 'message': 'Session expired. Please log in again.'}), 401
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_logged_in' not in session:
            if request.is_json or request.path.startswith('/api') or request.path == '/lookup-customer':
                return jsonify({'success': False, 'message': 'Session expired. Please log in again.'}), 401
            return redirect(url_for('login', next=request.url))
        if session.get('role') != 'admin':
            if request.is_json or request.path.startswith('/api') or request.path == '/lookup-customer':
                return jsonify({'success': False, 'message': 'Admin access required.'}), 403
            flash('Access denied. Admin privileges required.', 'error')
            return redirect(url_for('claim_status'))
        return f(*args, **kwargs)
    return decorated_function

def customercare_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_logged_in' not in session:
            if request.is_json or request.path.startswith('/api') or request.path == '/lookup-customer':
                return jsonify({'success': False, 'message': 'Session expired. Please log in again.'}), 401
            return redirect(url_for('login', next=request.url))
        if session.get('role') != 'customercare':
            if request.is_json or request.path.startswith('/api') or request.path == '/lookup-customer':
                return jsonify({'success': False, 'message': 'Access denied.'}), 403
            flash('Access denied.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def is_safe_url(target):
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, target))
    return test_url.scheme in ('http', 'https') and ref_url.netloc == test_url.netloc

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute; 30 per hour")
def login():
    if request.method == 'POST':
        # Verify CSRF Token
        session_token = session.get('csrf_token')
        form_token = request.form.get('csrf_token')
        if not session_token or form_token != session_token:
            flash('Invalid request session. Please try again.', 'error')
            return redirect(url_for('login'))

        # --- INPUT VALIDATION ---
        try:
            username = validate_username(request.form.get('username', ''))
        except ValueError as e:
            logging.warning(f"[LOGIN_INVALID_INPUT] {e} from IP {get_remote_address()}")
            flash('Invalid username or password', 'error')
            session['csrf_token'] = secrets.token_hex(32)
            return render_template('login.html')

        password = request.form.get('password', '')
        if not password or len(password) > 256:
            flash('Invalid username or password', 'error')
            session['csrf_token'] = secrets.token_hex(32)
            return render_template('login.html')
        # --- END VALIDATION ---

        user = USERS.get(username)
        if user and check_password_hash(user['password_hash'], password):
            session.permanent = True
            session.clear() # clear before setting new session to prevent fixation
            session['user_logged_in'] = True
            session['username'] = username
            session['role'] = user['role']
            session['display_name'] = user['display']
            flash('Login successful!', 'success')
            
            logging.info(f"[LOGIN_SUCCESS] User '{username}' successfully logged in from IP {get_remote_address()}")

            # Redirect based on role
            next_page = request.args.get('next')
            if next_page and is_safe_url(next_page):
                return redirect(next_page)
            if user['role'] == 'customercare':
                return redirect(url_for('claim_status'))
            return redirect(url_for('dashboard'))
        else:
            logging.warning(f"[LOGIN_FAILURE] Failed login attempt for username '{username}' from IP {get_remote_address()}")
            flash('Invalid username or password', 'error')
    
    # Generate CSRF for GET
    session['csrf_token'] = secrets.token_hex(32)
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

PASSWORD_RESET_TOKENS = {}

@app.route('/forgot-password', methods=['GET', 'POST'])
@limiter.limit("5 per minute; 10 per hour")
def forgot_password():
    if request.method == 'POST':
        # --- INPUT VALIDATION ---
        try:
            username = validate_username(request.form.get('username', ''))
        except ValueError:
            # Don't reveal whether username is valid
            flash('If the username exists, a reset link has been sent to the registered email.', 'info')
            return redirect(url_for('login'))
        # --- END VALIDATION ---
        user = USERS.get(username)
        if user and user.get('email'):
            token = secrets.token_hex(32)
            expires_at = datetime.datetime.now() + datetime.timedelta(minutes=15)
            PASSWORD_RESET_TOKENS[token] = {'username': username, 'expires_at': expires_at}
            reset_url = url_for('reset_password', token=token, _external=True)
            try:
                msg = MIMEMultipart()
                msg["From"] = SENDER_EMAIL
                msg["To"] = user['email']
                msg["Subject"] = "OSG Portal - Password Reset"
                body = f"Click <a href='{reset_url}'>here</a> to reset your password. The link expires in 15 minutes."
                msg.attach(MIMEText(body, "html"))
                with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                    server.starttls()
                    server.login(SENDER_EMAIL, SENDER_PASSWORD)
                    server.sendmail(SENDER_EMAIL, user['email'], msg.as_string())
            except Exception as e:
                print(f"Failed to send reset email: {e}")
        
        flash('If the username exists, a reset link has been sent to the registered email.', 'info')
        return redirect(url_for('login'))
    return render_template('forgot_password.html')

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    # --- INPUT VALIDATION: token format ---
    try:
        token = validate_reset_token(token)
    except ValueError:
        flash('Invalid or expired reset token.', 'error')
        return redirect(url_for('login'))

    token_data = PASSWORD_RESET_TOKENS.get(token)
    if not token_data or datetime.datetime.now() > token_data['expires_at']:
        flash('Invalid or expired reset token.', 'error')
        return redirect(url_for('login'))
        
    if request.method == 'POST':
        raw_password = request.form.get('password', '')
        # --- INPUT VALIDATION: password complexity ---
        try:
            new_password = validate_password_strength(raw_password, field_name='New password')
        except ValueError as e:
            flash(str(e), 'error')
            return render_template('reset_password.html', token=token)
        if new_password:
            username = token_data['username']
            new_hash = generate_password_hash(new_password)
            USERS[username]['password_hash'] = new_hash
            
            env_file = os.path.join(BASE_DIR, '.env')
            try:
                with open(env_file, 'r') as f:
                    content = f.read()
                prefix = 'ADMIN_PASSWORD_HASH=' if username == 'admin' else 'CUSTOMERCARE_PASSWORD_HASH='
                lines = content.split('\n')
                new_lines = [f"{prefix}{new_hash}" if line.startswith(prefix) else line for line in lines]
                with open(env_file, 'w') as f:
                    f.write('\n'.join(new_lines))
            except Exception as e:
                print(f"Failed to update .env: {e}")
                
            del PASSWORD_RESET_TOKENS[token]
            flash('Password reset successful! Please login with your new password.', 'success')
            return redirect(url_for('login'))
            
    return render_template('reset_password.html', token=token)

# ----------------------
# DATA MODEL (Wrapper)
# ----------------------
class ClaimWrapper:
    """Wraps dictionary data from Google Sheet to provide object-like access for templates"""
    def __init__(self, data):
        self.data = data
    
    def get(self, key, default=None):
        return self.data.get(key, default)
    
    # Generic access
    def __getattr__(self, name):
        # Map pythonic names to Sheet Headers
        # If specific logic needed, add property
        return self.data.get(name, self.data.get(name.replace('_', ' ').title(), None))

    # Properties matching the old SQLAlchemy model for template compatibility
    @property
    def id(self): return self.data.get("Claim ID") # Use Claim ID as ID
    @property
    def claim_id(self): return self.data.get("Claim ID")
    @property
    def created_at(self): 
        # Parse date string
        d = self.data.get("Date") or self.data.get("date")
        if not d: return datetime.datetime.now()
        s = str(d).strip()
        
        # Try multiple date formats
        formats_to_try = [
            '%Y-%m-%d %H:%M:%S',  # 2025-12-17 10:30:00
            '%Y-%m-%d',           # 2025-12-17
            '%d-%m-%Y',           # 17-12-2025
            '%d/%m/%Y',           # 17/12/2025
            '%m/%d/%Y',           # 12/17/2025
            '%d %b %Y',           # 17 Dec 2025
        ]
        
        for fmt in formats_to_try:
            try:
                if '%H:%M:%S' in fmt:
                    return datetime.datetime.strptime(s[:19], fmt)
                else:
                    return datetime.datetime.strptime(s[:10], fmt)
            except (ValueError, TypeError):
                continue
        
        # If all parsing fails, return current time
        return datetime.datetime.now()

    @property
    def customer_name(self): return self.data.get("Customer Name")
    @property
    def mobile_no(self): return self.data.get("Mobile Number")
    @property
    def address(self): return self.data.get("Address")
    @property
    def invoice_no(self): return self.data.get("Invoice Number")
    @property
    def serial_no(self): return self.data.get("Serial Number")
    @property
    def sr_no(self): return self.data.get("SR No")
    @property
    def model(self): return self.data.get("Model")
    @property
    def osid(self): return self.data.get("OSID")
    @property
    def issue(self): return self.data.get("Issue")
    @property
    def branch(self): return self.data.get("Branch")
    
    # Workflow
    @property
    def follow_up_date(self): return self.data.get("Follow Up - Dates")
    @property
    def follow_up_notes(self): return self.data.get("Follow Up - Notes")
    @property
    def claim_settled_date(self): return self.data.get("Claim Settled Date")
    @property
    def remarks(self): return self.data.get("Remarks")
    @property
    def status(self): return self.data.get("Status")
    
    # Booleans (Sheet has "Yes"/"No" or True/False or "1"/"0")
    # Google Sheets API sends Python booleans → stored as "True"/"False" strings in DB
    def _bool(self, key):
        val = self.data.get(key, "")
        v = str(val).strip().lower()
        return v in ("yes", "true", "1")

    @property
    def repair_feedback_completed(self): return self._bool("Repair Feedback Completed (Yes/No)")

    @property
    def feedback_rating(self): 
        val = self.data.get("Feedback Rating")
        if val is None:
            for k, v in self.data.items():
                if str(k).strip().lower() == "feedback rating":
                    val = v
                    break
                    
        val = str(val or "0").replace("'", "").strip()
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return 0

    @property
    def cust_confirmation_pending(self): 
        return self._bool("Replacement: Confirmation Pending") or self._bool("Customer Confirmation")
    
    @property
    def approval_mail_received(self): 
        return self._bool("Replacement: OSG Approval") or self._bool("Approval Mail Received From Onsitego (Yes/No)")
    
    @property
    def mail_sent_to_store(self):
        for key in ["Replacement: Mail to Store", "Mail Sent To Store (Yes/No)"]:
            if self._bool(key):
                return True
        # Fuzzy fallback: any key containing "mail" and "store"
        for k, v in self.data.items():
            if isinstance(k, str) and "mail" in k.lower() and "store" in k.lower():
                if str(v).strip().lower() in ("yes", "true", "1"):
                    return True
        return False
    
    @property
    def invoice_generated(self): 
        return self._bool("Replacement: Invoice Generated") or self._bool("Invoice Generated (Yes/No)")
    
    @property
    def invoice_sent_osg(self): 
        return self._bool("Replacement: Invoice Sent to OSG") or self._bool("Invoice Sent To Onsitego (Yes/No)")
    
    @property
    def settlement_mail_accounts(self):
        # Check all known column name variants
        for key in [
            "Replacement: Settlement Mail to Accounts",
            "Settlement Mail to Accounts(Yes/No)",
            "Settlement Mail to Accounts (Yes/No)",
        ]:
            if self._bool(key):
                return True
        # Fuzzy fallback: scan ALL data keys that relate to settlement mail
        for k, v in self.data.items():
            if not isinstance(k, str):
                continue
            kl = k.lower()
            if "settlement" in kl and "mail" in kl and "account" in kl:
                if str(v).strip().lower() in ("yes", "true", "1"):
                    return True
        return False
        
    @property
    def settled_with_accounts(self): 
        return self._bool("Replacement: Settled with Accounts") or self._bool("Settled With Accounts (Yes/No)")

    # --- Date properties for replacement workflow steps ---
    @property
    def approval_mail_date(self): return self.data.get("Approval Mail Received Date")

    @property
    def mail_sent_to_store_date(self): return self.data.get("Mail Sent To Store Date")

    @property
    def invoice_generated_date(self):
        # Check all possible column names the sheet may use
        return (self.data.get("Invoice Generated Date") or
                self.data.get("Invoice Generated Date (Yes/No)") or
                self.data.get("Replacement: Invoice Generated Date") or
                None)

    @property
    def invoice_sent_osg_date(self): return self.data.get("Invoice Sent To Onsitego Date")
    
    @property
    def complete(self):
        """A claim is complete if marked complete OR if status is Repair Completed/Closed OR all replacement workflow steps are done"""
        status = (self.status or "").strip().lower()
        
        # Explicitly exclude active statuses from being complete
        if status in ["submitted", "registered", "follow up"]:
            return False

        # Check complete checkbox
        if self._bool("Complete") or self._bool("Complete (Yes/No)"):
            return True
        
        # Also consider certain statuses as non-pending (resolved, rejected, cancelled, or on-call)
        if status in ["repair completed", "closed", "rejected", "no issue/oncall resolution",
                      "no issue", "oncall resolution", "cancelled"]:
            return True
        
        # Check if all replacement workflow steps are completed
        if "replacement" in status and "approved" in status:
            # A replacement claim is complete if all steps are done OR if mail is sent to store
            if self.mail_sent_to_store:
                return True
                
            all_steps_done = (
                self.cust_confirmation_pending and
                self.approval_mail_received and
                self.mail_sent_to_store and
                self.invoice_generated and
                self.invoice_sent_osg and
                self.settlement_mail_accounts and
                self.settled_with_accounts
            )
            if all_steps_done:
                return True
            
        return False

    @property
    def assigned_staff(self): return self.data.get("Assigned Staff")
    
    @property
    def tat(self):
        """Calculate TAT (Turnaround Time) in days"""
        # Return sheet value if it exists
        sheet_tat = self.data.get("Settled Time (TAT)")
        if sheet_tat and str(sheet_tat).strip() and str(sheet_tat) != 'nan':
            try:
                return int(float(sheet_tat))
            except (ValueError, TypeError, AttributeError):
                pass
        
        # Otherwise calculate it
        s_date = self.data.get("Date") or self.data.get("date") or self.data.get("Submitted Date")
        if self.claim_settled_date and s_date:
            try:
                submitted = datetime.datetime.strptime(str(s_date).split()[0], '%Y-%m-%d')
                settled = datetime.datetime.strptime(str(self.claim_settled_date).split()[0], '%Y-%m-%d')
                return (settled - submitted).days
            except Exception as e:
                return None
        return None
    
# ----------------------
# HELPER FUNCTIONS
# ----------------------
CLAIMS_CACHE = {
    'data': [],
    'last_updated': 0
}
CACHE_DURATION = 300  # 5 minute cache - balances freshness vs performance

def get_ist_now():
    return datetime.datetime.now(pytz.timezone('Asia/Kolkata'))

def invalidate_cache():
    global CLAIMS_CACHE
    print("Invalidating Cache...")
    CLAIMS_CACHE['last_updated'] = 0

def fetch_claims_from_db(force_refresh=False):
    global CLAIMS_CACHE
    import time
    
    current_time = time.time()
    if not force_refresh and (current_time - CLAIMS_CACHE['last_updated'] < CACHE_DURATION) and CLAIMS_CACHE['data']:
        print("Using Cached Data")
        return CLAIMS_CACHE['data']

    if not PG_AVAILABLE:
        logging.warning("[DB] PostgreSQL not configured, returning empty list.")
        return []
        
    try:
        logging.info("[DB] Fetching Fresh Data from PostgreSQL...")
        raw_rows = fetch_claims_from_postgres()
        # raw_rows can be empty if DB has 0 claims
        if raw_rows is not None:
            claims = [ClaimWrapper(d) for d in raw_rows]
            sorted_claims = sorted(claims, key=lambda x: x.created_at, reverse=True)
            
            # Update Cache
            CLAIMS_CACHE['data'] = sorted_claims
            CLAIMS_CACHE['last_updated'] = current_time
            
            return sorted_claims
    except Exception as e:
        logging.error(f"[DB] Fetch Error: {e}")
        
    if CLAIMS_CACHE['data']:
        print("Fetch failed, returning stale cache")
        return CLAIMS_CACHE['data']
    return []
    try:
        logging.info("[PG_FALLBACK] Attempting to load claims from PostgreSQL...")
        raw_rows = fetch_claims_from_postgres()
        if raw_rows:
            claims = [ClaimWrapper(d) for d in raw_rows]
            sorted_claims = sorted(claims, key=lambda x: x.created_at, reverse=True)
            logging.info(f"[PG_FALLBACK] Loaded {len(sorted_claims)} claims from PostgreSQL.")
            # Populate cache so subsequent in-process requests are fast
            CLAIMS_CACHE['data'] = sorted_claims
            CLAIMS_CACHE['last_updated'] = time.time()
            return sorted_claims
    except Exception as e:
        logging.error(f"[PG_FALLBACK] PostgreSQL fallback failed: {e}")
    return []

# ----------------------
# ROUTES
# ----------------------
@app.route('/')
@admin_required
def dashboard():
    refresh = request.args.get('refresh') == 'true'
    claims = fetch_claims_from_db(force_refresh=refresh)
    
    total = len(claims)
    pending = len([c for c in claims if not c.complete])
    completed = len([c for c in claims if c.complete])
    
    # Calculate Avg TAT
    tat_values = [c.tat for c in claims if c.tat is not None and isinstance(c.tat, int)]
    avg_tat = round(sum(tat_values) / len(tat_values)) if tat_values else 0

    # Calculate Data for OSG Customer Complaint Report
    now = get_ist_now().replace(tzinfo=None)
    report_stats = {
        'pending': {'lt5': 0, 'gt5': 0, 'gt10': 0, 'total': 0},
        'completed': 0,
        'rejected': 0,
        'replacement_mail': {'lt5': 0, 'gt5': 0, 'gt10': 0, 'total': 0},
        'gst_invoice': {'lt5': 0, 'gt5': 0, 'gt10': 0, 'total': 0},
        'pending_settlement_osg': {'lt5': 0, 'gt5': 0, 'gt10': 0, 'total': 0},
        'settlement_mail_accounts': {'lt5': 0, 'gt5': 0, 'gt10': 0, 'total': 0},
        'settled_accounts': 0,
        'grand_total_status': 0,
        'grand_total_replacement': 0,
        'report_date': now.strftime('%d-%m-%Y')
    }

    def _parse_date(raw):
        if not raw or str(raw).strip() in ('', 'nan', 'None'): return None
        s = str(raw).strip()[:10]
        dt = None
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
            try:
                dt = datetime.datetime.strptime(s, fmt)
                break
            except (ValueError, TypeError): continue
        
        # If it's a future date, Google Sheets likely interpreted DD-MM as MM-DD
        if dt and (dt - now).days > 1:
            try:
                if '-' in s and len(s.split('-')[0]) == 4:
                    dt = datetime.datetime.strptime(s, '%Y-%d-%m')
            except (ValueError, TypeError): pass
        return dt

    for c in claims:
        # Age for STATUS section: today − submitted date
        age = (now - c.created_at.replace(tzinfo=None)).days if c.created_at else 0

        # Age for REPLACEMENT section: today − claim_settled_date (if available), else today − submitted date
        settled_date_raw = c.claim_settled_date
        repl_age = age  # default fallback
        if settled_date_raw and str(settled_date_raw).strip() not in ('', 'nan', 'None'):
            try:
                settled_dt = datetime.datetime.strptime(str(settled_date_raw).strip()[:10], '%Y-%m-%d')
                repl_age = (now - settled_dt).days
            except Exception:
                try:
                    settled_dt = datetime.datetime.strptime(str(settled_date_raw).strip()[:10], '%d-%m-%Y')
                    repl_age = (now - settled_dt).days
                except Exception:
                    repl_age = age  # keep submitted-date age if parsing fails

        status = (c.status or "").strip().lower()
        is_replacement_claim = "replacement" in status

        # STATUS column logic (uses submitted-date age) — original correct logic
        if status == "rejected":
            report_stats['rejected'] += 1
            report_stats['grand_total_status'] += 1
        elif c.complete or status in ["repair completed", "closed", "no issue/oncall resolution",
                                       "no issue", "oncall resolution", "cancelled"]:
            report_stats['completed'] += 1
            report_stats['grand_total_status'] += 1
        else:
            report_stats['pending']['total'] += 1
            report_stats['grand_total_status'] += 1
            if age <= 5:
                report_stats['pending']['lt5'] += 1
            elif age <= 10:
                report_stats['pending']['gt5'] += 1
            else:
                report_stats['pending']['gt10'] += 1

        # REPLACEMENT column logic — waterfall: place claim at its CURRENT highest completed step
        # Aging buckets use claim_settled_date-based age (repl_age)
        if is_replacement_claim or c.mail_sent_to_store:
            if c.settled_with_accounts:
                report_stats['settled_accounts'] += 1
                report_stats['grand_total_replacement'] += 1
            elif c.settlement_mail_accounts:
                report_stats['settlement_mail_accounts']['total'] += 1
                report_stats['grand_total_replacement'] += 1
                if repl_age <= 5: report_stats['settlement_mail_accounts']['lt5'] += 1
                elif repl_age <= 10: report_stats['settlement_mail_accounts']['gt5'] += 1
                else: report_stats['settlement_mail_accounts']['gt10'] += 1
            elif c.invoice_sent_osg:
                report_stats['pending_settlement_osg']['total'] += 1
                report_stats['grand_total_replacement'] += 1
                osg_dt = _parse_date(c.invoice_sent_osg_date)
                if osg_dt:
                    osg_age = max(0, (now - osg_dt).days)
                else:
                    inv_gen_dt = _parse_date(c.invoice_generated_date)
                    if inv_gen_dt:
                        osg_age = max(0, (now - inv_gen_dt).days)
                    else:
                        osg_age = max(0, repl_age)
                if osg_age <= 5: report_stats['pending_settlement_osg']['lt5'] += 1
                elif osg_age <= 10: report_stats['pending_settlement_osg']['gt5'] += 1
                else: report_stats['pending_settlement_osg']['gt10'] += 1
            elif c.invoice_generated:
                report_stats['gst_invoice']['total'] += 1
                report_stats['grand_total_replacement'] += 1
                # Age from Invoice Generated Date (when GST invoice was billed)
                # Fallback chain: Invoice Generated Date → Mail Sent To Store Date → repl_age
                inv_gen_dt = _parse_date(c.invoice_generated_date)
                if inv_gen_dt:
                    gst_age = max(0, (now - inv_gen_dt).days)
                else:
                    store_dt = _parse_date(c.mail_sent_to_store_date)
                    if store_dt:
                        gst_age = max(0, (now - store_dt).days)
                    else:
                        gst_age = max(0, repl_age)

                if gst_age <= 5: report_stats['gst_invoice']['lt5'] += 1
                elif gst_age <= 10: report_stats['gst_invoice']['gt5'] += 1
                else: report_stats['gst_invoice']['gt10'] += 1
            else:
                # mail_sent_to_store step OR replacement approved with no checkboxes yet
                report_stats['replacement_mail']['total'] += 1
                report_stats['grand_total_replacement'] += 1
                if repl_age <= 5:
                    report_stats['replacement_mail']['lt5'] += 1
                elif repl_age <= 10:
                    report_stats['replacement_mail']['gt5'] += 1
                else:
                    report_stats['replacement_mail']['gt10'] += 1

    return render_template('dashboard.html', claims=claims, total=total, pending=pending, completed=completed, avg_tat=avg_tat, report_stats=report_stats)

@app.route('/health')
@app.route('/api/health')
def health_check():
    return jsonify({"status": "healthy", "timestamp": datetime.datetime.now().isoformat()})

@app.route('/favicon.ico')
def favicon():
    # Return no content to avoid 404/500 errors from browser favicon requests
    return '', 204

@app.route('/download-report')
@admin_required
def download_report():
    """Generate and download the OSG Customer Complaint Report as a styled Excel file."""
    import xlsxwriter

    # ── Rebuild the same report_stats as in the dashboard route ──
    claims = fetch_claims_from_db()
    now = get_ist_now().replace(tzinfo=None)

    report_stats = {
        'pending': {'lt5': 0, 'gt5': 0, 'gt10': 0, 'total': 0},
        'completed': 0,
        'rejected': 0,
        'replacement_mail': {'lt5': 0, 'gt5': 0, 'gt10': 0, 'total': 0},
        'gst_invoice': 0,
        'pending_settlement_osg': 0,
        'settlement_mail_accounts': 0,
        'settled_accounts': 0,
        'grand_total_status': 0,
        'grand_total_replacement': 0,
        'report_date': now.strftime('%d-%m-%Y')
    }

    for c in claims:
        age = (now - c.created_at.replace(tzinfo=None)).days if c.created_at else 0
        settled_date_raw = c.claim_settled_date
        repl_age = age
        if settled_date_raw and str(settled_date_raw).strip() not in ('', 'nan', 'None'):
            try:
                settled_dt = datetime.datetime.strptime(str(settled_date_raw).strip()[:10], '%Y-%m-%d')
                repl_age = (now - settled_dt).days
            except Exception:
                try:
                    settled_dt = datetime.datetime.strptime(str(settled_date_raw).strip()[:10], '%d-%m-%Y')
                    repl_age = (now - settled_dt).days
                except Exception:
                    repl_age = age

        status = (c.status or "").strip().lower()
        is_replacement_claim = "replacement" in status

        # STATUS column logic — original correct logic
        if status == "rejected":
            report_stats['rejected'] += 1
            report_stats['grand_total_status'] += 1
        elif c.complete or status in ["repair completed", "closed", "no issue/oncall resolution",
                                       "no issue", "oncall resolution", "cancelled"]:
            report_stats['completed'] += 1
            report_stats['grand_total_status'] += 1
        else:
            report_stats['pending']['total'] += 1
            report_stats['grand_total_status'] += 1
            if age <= 5:
                report_stats['pending']['lt5'] += 1
            elif age <= 10:
                report_stats['pending']['gt5'] += 1
            else:
                report_stats['pending']['gt10'] += 1

        if is_replacement_claim or c.mail_sent_to_store:
            if c.settled_with_accounts:
                report_stats['settled_accounts'] += 1
                report_stats['grand_total_replacement'] += 1
            elif c.settlement_mail_accounts:
                report_stats['settlement_mail_accounts'] += 1
                report_stats['grand_total_replacement'] += 1
            elif c.invoice_sent_osg:
                report_stats['pending_settlement_osg'] += 1
                report_stats['grand_total_replacement'] += 1
            elif c.invoice_generated:
                report_stats['gst_invoice'] += 1
                report_stats['grand_total_replacement'] += 1
            else:
                report_stats['replacement_mail']['total'] += 1
                report_stats['grand_total_replacement'] += 1
                if repl_age <= 5:
                    report_stats['replacement_mail']['lt5'] += 1
                elif repl_age <= 10:
                    report_stats['replacement_mail']['gt5'] += 1
                else:
                    report_stats['replacement_mail']['gt10'] += 1

    # ── Build Excel in memory ──
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    ws = workbook.add_worksheet('OSG Complaint Report')

    # ── Define formats matching the HTML table colours ──
    def fmt(props):
        defaults = {
            'border': 1, 'border_color': '#000000',
            'align': 'center', 'valign': 'vcenter',
            'font_name': 'Arial', 'font_size': 11,
        }
        defaults.update(props)
        return workbook.add_format(defaults)

    fmt_title   = fmt({'bold': True, 'font_size': 13, 'bg_color': '#FFFF00'})                     # yellow header
    fmt_blue    = fmt({'bold': True, 'bg_color': '#9BC2E6'})                                       # light-blue header cells
    fmt_green   = fmt({'bold': True, 'bg_color': '#C6E0B4'})                                       # green <5
    fmt_orange  = fmt({'bold': True, 'bg_color': '#FFC000'})                                       # dark-orange >5
    fmt_red     = fmt({'bold': True, 'bg_color': '#FF0000', 'font_color': '#FFFFFF'})              # red >10
    fmt_peach   = fmt({'bg_color': '#F8CBAD'})                                                     # peach pending settlement
    fmt_normal  = fmt({})                                                                           # plain cell
    fmt_bold    = fmt({'bold': True})                                                               # bold plain
    fmt_blue_bg = fmt({'bold': True, 'bg_color': '#9BC2E6'})                                       # grand total rows

    # ── Column widths ──
    ws.set_column(0, 0, 22)   # STATUS
    ws.set_column(1, 3, 8)    # <5 >5 >10
    ws.set_column(4, 4, 10)   # TOTAL
    ws.set_column(5, 5, 36)   # REPLACEMENT label
    ws.set_column(6, 8, 8)    # <5 >5 >10
    ws.set_column(9, 9, 10)   # TOTAL
    ws.set_row(0, 22)          # title row height
    for r in range(1, 10):
        ws.set_row(r, 18)

    # ── Row 0: Title (A1:J1) ──
    ws.merge_range('A1:J1', f'OSG CUSTOMER COMPLAINT REPORT {report_stats["report_date"]}', fmt_title)

    # ── Row 1: Column group headers ──
    ws.merge_range('A2:A3', 'STATUS',      fmt_blue)
    ws.merge_range('B2:E2', 'COUNT',       fmt_blue)
    ws.merge_range('F2:F3', 'REPLACEMENT', fmt_blue)
    ws.merge_range('G2:J2', 'COUNT',       fmt_blue)

    # ── Row 2: Sub-headers <5 >5 >10 TOTAL ──
    ws.write('B3', '<5',    fmt_green)
    ws.write('C3', '>5',    fmt_orange)
    ws.write('D3', '>10',   fmt_red)
    ws.write('E3', 'TOTAL', fmt_blue)
    ws.write('G3', '<5',    fmt_green)
    ws.write('H3', '>5',    fmt_orange)
    ws.write('I3', '>10',   fmt_red)
    ws.write('J3', 'TOTAL', fmt_blue)

    # ── Data rows (rows 3-6, i.e. Excel rows 4-7) ──
    # PENDING row
    ws.write('A4', 'PENDING',                                     fmt_bold)
    ws.write('B4', report_stats['pending']['lt5'],                 fmt_normal)
    ws.write('C4', report_stats['pending']['gt5'],                 fmt_normal)
    ws.write('D4', report_stats['pending']['gt10'],                fmt_normal)
    ws.write('E4', report_stats['pending']['total'],               fmt_bold)
    ws.write('F4', 'REPLACEMENT MAIL SENT TO STORE',              fmt_normal)
    ws.write('G4', report_stats['replacement_mail']['lt5'],        fmt_normal)
    ws.write('H4', report_stats['replacement_mail']['gt5'],        fmt_normal)
    ws.write('I4', report_stats['replacement_mail']['gt10'],       fmt_normal)
    ws.write('J4', report_stats['replacement_mail']['total'],      fmt_bold)

    # COMPLETED row
    ws.write('A5', 'COMPLETED',                                    fmt_bold)
    ws.merge_range('B5:E5', report_stats['completed'],             fmt_normal)
    ws.write('F5', 'GST INVOICE BILLED',                          fmt_normal)
    ws.merge_range('G5:J5', report_stats['gst_invoice'],           fmt_normal)

    # REJECTED rows (merged A:E across 3 rows)
    ws.merge_range('A6:A8', 'REJECTED',                           fmt_bold)
    ws.merge_range('B6:E8', report_stats['rejected'],             fmt_normal)
    ws.write('F6', 'PENDING SETTLEMENT FROM OSG',                 fmt_peach)
    ws.merge_range('G6:J6', report_stats['pending_settlement_osg'], fmt_normal)
    ws.write('F7', 'SETTLEMENT MAIL SENT TO myG ACCOUNTS',        fmt_normal)
    ws.merge_range('G7:J7', report_stats['settlement_mail_accounts'], fmt_normal)
    ws.write('F8', 'SETTLED WITH ACCOUNTS',                       fmt_normal)
    ws.merge_range('G8:J8', report_stats['settled_accounts'],     fmt_normal)

    # GRAND TOTAL row
    ws.write('A9', 'GRAND TOTAL',                                 fmt_blue_bg)
    ws.merge_range('B9:E9', report_stats['grand_total_status'],   fmt_blue_bg)
    ws.write('F9', 'GRAND TOTAL',                                 fmt_blue_bg)
    ws.merge_range('G9:J9', report_stats['grand_total_replacement'], fmt_blue_bg)

    workbook.close()
    output.seek(0)

    filename = f"OSG_Complaint_Report_{report_stats['report_date']}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ──────────────────────────────────────────────────────────────────────────
# CUSTOMER LOOKUP — PostgreSQL (osid_data table)
# ──────────────────────────────────────────────────────────────────────────
def lookup_customer_from_db(mobile: str) -> dict:
    """
    Query the osid_data table for all products linked to a mobile number.
    Returns {'name': str, 'products': [...]} or None if not found.
    """
    try:
        import psycopg2
        import psycopg2.extras
        db_url = os.environ.get('DATABASE_URL')
        if not db_url:
            logging.warning('[LOOKUP] DATABASE_URL not set.')
            return None

        conn = psycopg2.connect(db_url)
        try:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("""
                    SELECT customer, invoice_no, model, serial_no, osid, store_name
                    FROM osid_data
                    WHERE TRIM(mobile_no) = %s
                    ORDER BY date DESC
                """, (mobile,))
                rows = cur.fetchall()
        finally:
            conn.close()

        if not rows:
            return None

        name = rows[0]['customer'] or 'Unknown'
        products = [
            {
                'invoice': r['invoice_no'] or '',
                'model':   r['model']      or '',
                'serial':  r['serial_no']  or '',
                'osid':    r['osid']       or '',
                'branch':  r['store_name'] or 'Main Branch',
            }
            for r in rows
        ]
        return {'name': name, 'products': products}

    except Exception as e:
        logging.error(f'[LOOKUP] DB lookup error: {e}')
        return None


# Stubs — kept so any remaining internal references don't break
def load_excel_data():
    return {}

def col_lookup(df, variations):
    for v in variations:
        if v in (df.columns if hasattr(df, 'columns') else []):
            return v
    return None


@app.route('/lookup-customer', methods=['POST'])
@login_required
def lookup_customer():
    data = request.json
    if not data:
        return jsonify({"success": False, "message": "Invalid request body."})

    # --- INPUT VALIDATION ---
    try:
        mobile = validate_mobile(data.get('mobile', ''))
    except ValueError as e:
        return jsonify({"success": False, "message": str(e)})

    # Query PostgreSQL osid_data table directly
    customer_data = lookup_customer_from_db(mobile)

    if customer_data:
        logging.info(f"[LOOKUP] Found {len(customer_data['products'])} product(s) for mobile {mobile}")
        return jsonify({
            "success": True,
            "customer_name": customer_data['name'],
            "products": customer_data['products']
        })
    else:
        logging.info(f"[LOOKUP] No records found for mobile: {mobile}")
        return jsonify({"success": False, "message": "No customer found with this mobile number."})

def send_email_notification(claim_data, files=None):
    try:
        msg = MIMEMultipart()
        msg["From"] = SENDER_EMAIL
        msg["To"] = TARGET_EMAIL
        msg["Cc"] = ", ".join(CC_EMAILS)
        msg["Subject"] = f"🛡️ Warranty Claim Submission – OSID: {claim_data.get('osid', 'N/A')} – {claim_data.get('customer_name', 'Unknown')}"
        
        body = f"""
        <html><body>
        <div style="font-family: Arial, sans-serif; max-width: 800px; margin: 0 auto;">
            <div style="background: linear-gradient(135deg, #2E86C1 0%, #5DADE2 100%); color: white; padding: 20px; text-align: center; border-radius: 10px 10px 0 0;">
                <h2 style="margin: 0;">🛡️ Warranty Claim Submission</h2>
                <p style="margin: 5px 0 0 0;">New claim received from customer</p>
            </div>
            <div style="background: #f8f9fa; padding: 20px; border-radius: 0 0 10px 10px;">
                <p>Dear SR team,</p>
                <p>We have received a warranty claim for the products purchased by our customer. Please find the details below:</p>
                <div style="background: white; padding: 15px; border-radius: 8px; margin: 12px 0; border-left: 4px solid #2E86C1;">
                    <h3 style="color: #2E86C1; margin-top: 0;">👤 Customer Information</h3>
                    <p><strong>Name:</strong> {claim_data.get('customer_name')}<br>
                    <strong>Mobile No:</strong> {claim_data.get('mobile_no')}<br>
                    <strong>Address:</strong> {claim_data.get('address')}</p>
                </div>
                <div style="background: white; padding: 15px; border-radius: 8px; margin: 12px 0; border-left: 4px solid #28A745;">
                    <h3 style="color: #28A745; margin-top: 0;">📦 Product Details & Issue</h3>
                    <p><strong>Model:</strong> {claim_data.get('model')}<br>
                    <strong>Serial:</strong> {claim_data.get('serial_no')}<br>
                    <strong>OSID:</strong> {claim_data.get('osid')}<br>
                    <strong>Invoice:</strong> {claim_data.get('invoice_no')}<br>
                    <strong>Issue:</strong> {claim_data.get('issue')}</p>
                </div>
                <div style="background: #e7f3ff; padding: 12px; border-radius: 8px; margin: 12px 0;">
                    <p><strong>📅 Submitted:</strong> {get_ist_now().strftime('%Y-%m-%d %H:%M:%S IST')}</p>
                </div>
            </div>
        </div>
        </body></html>
        """
        msg.attach(MIMEText(body, "html"))

        if files:
            for f in files:
                try:
                    with open(f, "rb") as fil:
                        part = MIMEApplication(fil.read(), Name=os.path.basename(f))
                        part['Content-Disposition'] = f'attachment; filename="{os.path.basename(f)}"'
                        msg.attach(part)
                except Exception as e:
                    print(f"Failed to attach file: {e}")

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.sendmail(SENDER_EMAIL, [TARGET_EMAIL] + CC_EMAILS, msg.as_string())
        return True
    except Exception as e:
        print(f"Email error: {e}")
        return False

# ----------------------
# ROUTES
# ----------------------

@app.route('/submit-claim', methods=['GET', 'POST'])
@admin_required
def submit_claim():
    if request.method == 'GET':
        return render_template('submit.html')
    
    # Handle POST
    try:
        data = request.form

        # --- INPUT VALIDATION ---
        try:
            customer_name = sanitize_text(data.get('customer_name', ''), max_len=MAX_TEXT_SHORT, field_name='Customer name', sheet_destined=True)
            mobile        = validate_mobile(data.get('mobile', ''))
            address       = sanitize_text(data.get('address', ''), max_len=MAX_TEXT_LONG, field_name='Address', allow_newlines=True, sheet_destined=True)
            if not customer_name:
                return jsonify({"success": False, "message": "Customer name is required."})
        except ValueError as e:
            return jsonify({"success": False, "message": str(e)})
        # --- END VALIDATION ---

        claims_json = data.get('claims_data')
        if not claims_json:
            if data.get('selected_product'):
                import json
                prod = json.loads(data.get('selected_product'))
                prod['issue'] = data.get('issue')
                prod['file_key'] = 'files'
                claims_json = json.dumps([prod])
            else:
                return jsonify({"success": False, "message": "No claims data received"})

        import json
        claims_list = json.loads(claims_json)

        # Enforce maximum claims per submission
        if not isinstance(claims_list, list) or len(claims_list) == 0:
            return jsonify({"success": False, "message": "claims_data must be a non-empty list."})
        if len(claims_list) > 10:
            return jsonify({"success": False, "message": "Maximum 10 claims allowed per submission."})

        results = []

        # Ensure upload folder exists
        if not os.path.exists(UPLOAD_FOLDER):
            os.makedirs(UPLOAD_FOLDER)

        for idx, item in enumerate(claims_list):
            if not isinstance(item, dict):
                return jsonify({"success": False, "message": f"claims_data[{idx}] must be an object."})

            # Sanitize claim item fields
            try:
                item_issue   = sanitize_text(item.get('issue', ''),   max_len=MAX_TEXT_DEFAULT, field_name=f'Issue[{idx}]',          sheet_destined=True)
                item_model   = sanitize_text(item.get('model', ''),   max_len=MAX_TEXT_SHORT,   field_name=f'Model[{idx}]',           sheet_destined=True)
                item_serial  = sanitize_text(item.get('serial', ''),  max_len=MAX_TEXT_SHORT,   field_name=f'Serial number[{idx}]',   sheet_destined=True)
                item_osid    = sanitize_text(item.get('osid', ''),    max_len=MAX_TEXT_SHORT,   field_name=f'OSID[{idx}]',            sheet_destined=True)
                item_invoice = sanitize_text(item.get('invoice', ''), max_len=MAX_TEXT_SHORT,   field_name=f'Invoice number[{idx}]',  sheet_destined=True)
                item_branch  = sanitize_text(item.get('branch', 'Main Branch'), max_len=MAX_TEXT_SHORT, field_name=f'Branch[{idx}]', sheet_destined=True)
                item_fkey    = sanitize_text(item.get('file_key', ''), max_len=50,              field_name=f'file_key[{idx}]')
            except ValueError as e:
                return jsonify({"success": False, "message": str(e)})

            # Files – validate each upload before saving
            uploaded_files = request.files.getlist(item_fkey) if item_fkey else []
            if not uploaded_files and item_fkey == 'files':
                uploaded_files = request.files.getlist('files')

            saved_paths = []
            for f in uploaded_files:
                if f and f.filename:
                    try:
                        validate_file_upload(f, field_name=f'Attachment[{idx}]')
                    except ValueError as e:
                        return jsonify({"success": False, "message": str(e)})
                    fn = secure_filename(f"{int(time.time())}_{idx}_{f.filename}")
                    path = os.path.join(UPLOAD_FOLDER, fn)
                    f.save(path)
                    saved_paths.append(path)

            # Lookup missing fields from osid_data
            if not item_osid or not item_invoice or not item_serial:
                try:
                    from services.pg_sync import _get_connection
                    import psycopg2.extras
                    conn = _get_connection()
                    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                        if mobile:
                            cur.execute("SELECT osid, invoice_no, serial_no FROM osid_data WHERE mobile_no = %s ORDER BY id DESC LIMIT 1;", (mobile,))
                            row = cur.fetchone()
                            if row:
                                item_osid = item_osid or (row.get('osid') or "")
                                item_invoice = item_invoice or (row.get('invoice_no') or "")
                                item_serial = item_serial or (row.get('serial_no') or "")
                except Exception as e:
                    print(f"Error lookup osid_data on submit: {e}")

            # Build claim object with sanitized values
            unique_suffix = int(time.time()) + idx
            new_claim = {
                "Claim ID": f"CLM-{unique_suffix}",
                "Date": get_ist_now().strftime('%Y-%m-%d'),
                "Customer Name": customer_name,
                "Mobile Number": mobile,
                "Address": address,
                "Product": item_model,
                "Invoice Number": item_invoice,
                "Serial Number": item_serial,
                "Model": item_model,
                "OSID": item_osid,
                "Branch": item_branch,
                "Issue": item_issue,
                "Status": "Submitted"
            }

            print(f"Syncing Claim {idx+1}/{len(claims_list)}: {new_claim['Claim ID']}")
            sync_to_database_dict(new_claim)

            # Push to Google Sheets in background (only basic complaint columns)
            web_app_url = os.environ.get("WEB_APP_URL")
            if web_app_url:
                import threading, requests
                def _push_to_sheet(url, data):
                    # Column names MUST match the actual Google Sheet headers exactly
                    sheet_payload = {
                        "Claim ID":       data.get("Claim ID"),
                        "Submitted Date": data.get("Date"),          # Sheet uses "Submitted Date"
                        "Customer Name":  data.get("Customer Name"),
                        "Mobile":         data.get("Mobile Number"),  # Sheet uses "Mobile"
                        "Branch":         data.get("Branch"),
                        "Product":        data.get("Product"),
                        "Issue":          data.get("Issue"),
                        "Status":         data.get("Status"),
                        "OSID":           data.get("OSID"),
                        "Serial Number":  data.get("Serial Number"),
                        "Invoice Number": data.get("Invoice Number"),
                    }
                    # Remove None values
                    sheet_payload = {k: v for k, v in sheet_payload.items() if v is not None}
                    try:
                        print(f"[SHEET_PUSH] Pushing new claim to Google Sheets: {sheet_payload.get('Claim ID')}")
                        response = requests.post(url, json=sheet_payload, timeout=15)
                        print(f"[SHEET_PUSH] Response: {response.status_code} - {response.text}")
                    except Exception as e:
                        print(f"[SHEET_PUSH] Failed: {e}")
                threading.Thread(target=_push_to_sheet, args=(web_app_url, new_claim)).start()

            send_email_notification({
                "customer_name": customer_name,
                "mobile_no": mobile,
                "address": address,
                "model": item_model,
                "serial_no": item_serial,
                "osid": item_osid,
                "invoice_no": item_invoice,
                "issue": item_issue
            }, saved_paths)

            results.append(new_claim["Claim ID"])
            time.sleep(0.5)

        invalidate_cache()
        return jsonify({"success": True, "message": f"Successfully submitted {len(results)} claim(s)!"})

    except Exception as e:
        print(f"Submit Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)})

@app.route('/claim/<string:id>', methods=['GET'])
@login_required
def get_claim(id):
    # --- INPUT VALIDATION ---
    try:
        id = validate_claim_id(id, field_name='claim id')
    except ValueError:
        return jsonify({"error": "Invalid claim ID format."}), 400
    # --- END VALIDATION ---
    claims = fetch_claims_from_db()
    
    # Find claim by Claim ID (id) or numeric ID? User passes int ID earlier, but now strings CLM-...
    # Let's support both if possible or just filter by Claim ID
    
    found = None
    for c in claims:
        # Check against "Claim ID"
        if str(c.claim_id) == str(id):
            found = c
            break
            
    if not found:
        return jsonify({"error": "Not found"}), 404

    # Guard: found must be set (already checked above, but assert removed for safety)
    if found is None:
        return jsonify({"error": "Not found"}), 404

    # Convert Wrapper to dict for frontend
    # We need to map back to the keys JS expects
    
    # Helper to parse boolean values
    def parse_bool(val):
        if val is None or val == '':
            return False
        return str(val).strip().lower() in ['yes', 'true', '1']
    
    invoice_no = found.invoice_no or ""
    serial_no = found.serial_no or ""
    osid = found.osid or ""
    
    # Fetch from osid_data if missing
    if not invoice_no or not osid or not serial_no:
        try:
            from services.pg_sync import _get_connection
            import psycopg2.extras
            conn = _get_connection()
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                mobile = str(found.mobile_no or "").strip()
                if mobile:
                    cur.execute("SELECT osid, invoice_no, serial_no FROM osid_data WHERE mobile_no = %s ORDER BY id DESC LIMIT 1;", (mobile,))
                    row = cur.fetchone()
                    if row:
                        osid = osid or (row.get('osid') or "")
                        invoice_no = invoice_no or (row.get('invoice_no') or "")
                        serial_no = serial_no or (row.get('serial_no') or "")
        except Exception as e:
            print(f"Error fetching from osid_data: {e}")

    # ---------------------------------------------------------------------
    # Implicit Checkbox Logic
    # If status progressed to "Replacement Approved", assume prior steps are done.
    # ---------------------------------------------------------------------
    status_val = str(found.status or "").strip().lower()
    is_rep_approved = (status_val == "replacement approved")
    is_rep_closed = (status_val == "replacement closed")
    
    rep_conf = parse_bool(found.data.get("Customer Confirmation"))
    rep_osg_app = parse_bool(found.data.get("Approval Mail Received From Onsitego (Yes/No)"))
    rep_mail_store = parse_bool(found.data.get("Mail Sent To Store (Yes/No)"))
    
    if is_rep_approved or is_rep_closed:
        rep_conf = True
        rep_osg_app = True
        rep_mail_store = True

    return jsonify({
        "id": found.claim_id,
        "date": found.created_at.strftime('%Y-%m-%d'),
        "customer_name": found.customer_name,
        "mobile_no": found.mobile_no or "",
        "invoice_no": invoice_no,
        "serial_no": serial_no,
        "model": found.model or "",
        "issue": found.issue or "",
        "address": found.address or "",
        "claim_settled_date": found.claim_settled_date or "",
        "status": found.status,
        "follow_up_date": found.follow_up_date or "",
        "follow_up_notes": found.follow_up_notes or "",
        "remarks": found.remarks or "",
        "repair_feedback_completed": found.repair_feedback_completed,
        "feedback_rating": found.feedback_rating,
        
        # Replacement workflow fields (Columns O-T) - Use actual sheet column names
        "replacement_confirmation": rep_conf,
        "replacement_osg_approval": rep_osg_app,
        "replacement_mail_store": rep_mail_store,
        "replacement_invoice_gen": parse_bool(found.data.get("Invoice Generated (Yes/No)")),
        "replacement_invoice_sent": parse_bool(found.data.get("Invoice Sent To Onsitego (Yes/No)")),
        "settlement_mail_accounts": parse_bool(found.data.get("Settlement Mail to Accounts(Yes/No)")),
        "replacement_settled_accounts": parse_bool(found.data.get("Settled With Accounts (Yes/No)")),
        
        # Complete flag
        "complete": parse_bool(found.data.get("Complete (Yes/No)")),
        
        "tat": found.tat,
        "assigned_staff": found.assigned_staff or "",
        "sr_no": found.sr_no or "",
        "osid": osid
    })

@app.route('/update-claim/<string:id>', methods=['POST'])
@login_required
def update_claim(id):
    # --- INPUT VALIDATION: URL param ---
    try:
        id = validate_claim_id(id, field_name='claim id')
    except ValueError:
        return jsonify({'success': False, 'message': 'Invalid claim ID format.'}), 400

    data = request.json
    if not data or not isinstance(data, dict):
        return jsonify({'success': False, 'message': 'Invalid JSON body.'}), 400

    # --- INPUT VALIDATION: fields ---
    try:
        if 'status' in data:
            validate_status(data['status'])   # raises if not in whitelist
        for date_field in ('date', 'follow_up_date', 'approval_mail_date',
                           'mail_sent_to_store_date', 'invoice_generated_date',
                           'invoice_sent_osg_date', 'claim_settled_date'):
            if date_field in data:
                data[date_field] = validate_date_str(data[date_field], field_name=date_field)
        for text_field in ('follow_up_notes', 'remarks'):
            if text_field in data:
                data[text_field] = sanitize_text(data[text_field], max_len=MAX_TEXT_LONG,
                                                  field_name=text_field, allow_newlines=True,
                                                  sheet_destined=True)
        for short_field in ('assigned_staff', 'sr_no'):
            if short_field in data:
                data[short_field] = sanitize_text(data[short_field], max_len=MAX_TEXT_SHORT,
                                                   field_name=short_field, sheet_destined=True)
        if 'feedback_rating' in data:
            val_str = str(data['feedback_rating']).strip()
            if val_str in ['0', '', 'None', 'null', '0.0']:
                del data['feedback_rating']
            else:
                data['feedback_rating'] = validate_integer_range(data['feedback_rating'], 1, 5,
                                                                  field_name='feedback_rating')
        for bool_field in ('repair_feedback_completed', 'replacement_confirmation',
                           'replacement_osg_approval', 'replacement_mail_store',
                           'replacement_invoice_gen', 'replacement_invoice_sent',
                           'replacement_settlement_mail', 'replacement_settled_accounts', 'complete'):
            if bool_field in data:
                data[bool_field] = validate_boolean(data[bool_field], field_name=bool_field)
    except ValueError as e:
        logging.warning(f"[UPDATE_CLAIM_INVALID] {e} on claim {id} from IP {get_remote_address()}")
        return jsonify({'success': False, 'message': str(e)}), 400
    # --- END VALIDATION ---

    # Map JS keys back to Sheet Headers
    payload = {
        "Claim ID": id
    }

    if 'status' in data: payload["Status"] = data['status']
    if 'date' in data: payload["Date"] = data['date']
    if 'follow_up_notes' in data: payload["Follow Up - Notes"] = data['follow_up_notes']
    if 'remarks' in data: payload["Remarks"] = data['remarks']
    if 'assigned_staff' in data: payload["Assigned Staff"] = data['assigned_staff']
    if 'sr_no' in data and data['sr_no'].strip(): payload["SR No"] = data['sr_no']  # Only set SR No if non-empty

    if 'follow_up_date' in data: payload["Follow Up - Dates"] = data['follow_up_date']
    if 'approval_mail_date' in data: payload["Approval Mail Received Date"] = data['approval_mail_date']
    if 'mail_sent_to_store_date' in data: payload["Mail Sent To Store Date"] = data['mail_sent_to_store_date']
    if 'invoice_generated_date' in data: payload["Invoice Generated Date"] = data['invoice_generated_date']
    if 'invoice_sent_osg_date' in data: payload["Invoice Sent To Onsitego Date"] = data['invoice_sent_osg_date']
    if 'claim_settled_date' in data: payload["Claim Settled Date"] = data['claim_settled_date']

    def fmt_bool(val): return "Yes" if val else "No"
    
    if 'repair_feedback_completed' in data: payload["Repair Feedback Completed (Yes/No)"] = fmt_bool(data['repair_feedback_completed'])
    if 'feedback_rating' in data: payload["Feedback Rating"] = f"'{data['feedback_rating']}"
    
    # Find existing claim to check for existing dates
    all_claims = fetch_claims_from_db()
    existing_claim = next((c for c in all_claims if str(c.claim_id) == str(id)), None)
    
    # --- WHATSAPP NOTIFICATION LOGIC ---
    if 'status' in data:
        new_status = data['status'].strip().upper()
        existing_status = (existing_claim.status or "").strip().upper() if existing_claim else ""
        last_notified_status = (existing_claim.data.get("Last_Notified_Status") or "").strip().upper() if existing_claim else ""
        
        # Helper to check boolean fields for Replacement Workflow
        def is_checked(bool_key, db_key):
            if bool_key in data:
                return data[bool_key]
            if existing_claim and existing_claim.data:
                val = existing_claim.data.get(db_key)
                if val is not None and str(val).strip() != '':
                    return str(val).strip().lower() in ['yes', 'true', '1']
            return False

        repl_confirmed = is_checked('replacement_confirmation', 'Customer Confirmation')
        repl_osg = is_checked('replacement_osg_approval', 'Approval Mail Received From Onsitego (Yes/No)')
        repl_mail = is_checked('replacement_mail_store', 'Mail Sent To Store (Yes/No)')

        should_notify = False
        
        if new_status != existing_status and new_status != last_notified_status:
            if new_status in ["REGISTERED", "REPAIR COMPLETED", "REJECTED"]:
                should_notify = True
                
        if new_status == "REPLACEMENT APPROVED" and last_notified_status != "REPLACEMENT APPROVED":
            if repl_confirmed and repl_osg and repl_mail:
                should_notify = True

        if should_notify:
                # --- CUTOFF DATE GATE ---
                # Block notifications for claims registered before the cutoff date
                claim_registered_date = existing_claim.created_at if existing_claim else None
                if claim_registered_date:
                    # Strip timezone info for comparison if present
                    reg_date_naive = claim_registered_date.replace(tzinfo=None)
                    if reg_date_naive < WHATSAPP_CUTOFF_DATE:
                        logging.info(
                            f"[WHATSAPP_BLOCKED] Claim {id} registered on "
                            f"{reg_date_naive.strftime('%Y-%m-%d')} is before cutoff "
                            f"{WHATSAPP_CUTOFF_DATE.strftime('%Y-%m-%d')}. Skipping notification."
                        )
                    else:
                        # Registration date is on or after the cutoff — send message
                        from services.whatsapp_service import send_whatsapp_message
                        mobile = str(existing_claim.mobile_no) if existing_claim and existing_claim.mobile_no else data.get('mobile', '')
                        c_name = (existing_claim.customer_name if existing_claim and existing_claim.customer_name else data.get('customer_name', '')).strip()
                        if not c_name: c_name = "Customer"
                        
                        c_model = (existing_claim.model if existing_claim and existing_claim.model else data.get('model', '')).strip()
                        if not c_model: c_model = "your product"
                        
                        c_sr_no = (existing_claim.sr_no if existing_claim and existing_claim.sr_no else data.get('sr_no', '')).strip()
                        if not c_sr_no:
                            c_sr_no = id  # Fallback to Claim ID (e.g. CLM-123) if SR No is missing

                        if new_status == "REPAIR COMPLETED":
                            template_to_use = "myg_onsitego_repair_completed_main"
                            template_params = [c_name, c_sr_no]
                        elif new_status == "REPLACEMENT APPROVED":
                            template_to_use = "myg_onsitego_replacement_main"
                            template_params = [c_name, c_model]
                        elif new_status == "REJECTED":
                            template_to_use = "osg_clm_reject"
                            template_params = [c_name]
                        else:
                            template_to_use = "myg_onsitego_registered_main"
                            template_params = [c_name, c_model, c_sr_no]

                        resp = send_whatsapp_message(
                            mobile=mobile,
                            template_name=template_to_use,
                            params=template_params
                        )
                        print(f"WhatsApp Trigger Response ({new_status}):", resp)
                        
                        # DEBUG LOGGING FOR WHATSAPP
                        try:
                            with open("whatsapp_debug_log.txt", "a") as f:
                                f.write(f"--- WHATSAPP TRIGGER ---\n")
                                f.write(f"Status: {new_status}\n")
                                f.write(f"Template: {template_to_use}\n")
                                f.write(f"Params: {template_params}\n")
                                f.write(f"Response: {resp}\n\n")
                        except Exception:
                            pass
                        
                        # Only update Last_Notified_Status if message was successfully sent (not blocked and API returned 2xx)
                        if not resp.get("blocked") and resp.get("status_code") in [200, 201, 202]:
                            payload["Last_Notified_Status"] = new_status
                else:
                    logging.warning(f"[WHATSAPP_BLOCKED] Claim {id} has no registration date. Skipping notification.")
                # --- END CUTOFF DATE GATE ---

    # ------------------------------------

    import datetime
    import pytz
    
    # Get current time in IST
    ist = pytz.timezone('Asia/Kolkata')
    today_str = datetime.datetime.now(ist).strftime('%d-%m-%Y')

    def should_update_date(key_bool, existing_date_val):
        # Update date if: Checkbox is TRUE AND (Existing Date is Empty/None)
        is_checked = data.get(key_bool)
        if is_checked:
            if not existing_date_val or str(existing_date_val).strip() == '':
                return True
        return False

    # Replacement workflow fields (Columns O-T) - Use actual sheet column names
    if 'replacement_confirmation' in data: payload["Customer Confirmation"] = fmt_bool(data['replacement_confirmation'])
    
    # Auto-date logic for: Approval Mail
    if 'replacement_osg_approval' in data: 
        payload["Approval Mail Received From Onsitego (Yes/No)"] = fmt_bool(data['replacement_osg_approval'])
        # Check if we need to set date
        existing_date = existing_claim.approval_mail_date if existing_claim else None
        if should_update_date('replacement_osg_approval', existing_date):
             payload["Approval Mail Received Date"] = today_str

    # Auto-date logic for: Mail Sent To Store
    if 'replacement_mail_store' in data: 
        payload["Mail Sent To Store (Yes/No)"] = fmt_bool(data['replacement_mail_store'])
        existing_date = existing_claim.mail_sent_to_store_date if existing_claim else None
        if should_update_date('replacement_mail_store', existing_date):
             payload["Mail Sent To Store Date"] = today_str

    # Auto-date logic for: Invoice Generated
    if 'replacement_invoice_gen' in data: 
        payload["Invoice Generated (Yes/No)"] = fmt_bool(data['replacement_invoice_gen'])
        existing_date = existing_claim.invoice_generated_date if existing_claim else None
        if should_update_date('replacement_invoice_gen', existing_date):
             payload["Invoice Generated Date"] = today_str

    # Auto-date logic for: Invoice Sent To Onsitego
    if 'replacement_invoice_sent' in data: 
        payload["Invoice Sent To Onsitego (Yes/No)"] = fmt_bool(data['replacement_invoice_sent'])
        existing_date = existing_claim.invoice_sent_osg_date if existing_claim else None
        if should_update_date('replacement_invoice_sent', existing_date):
            payload["Invoice Sent To Onsitego Date"] = today_str

    # Auto-date logic for: Settlement Mail to Accounts
    if 'replacement_settlement_mail' in data: 
        payload["Settlement Mail to Accounts(Yes/No)"] = fmt_bool(data['replacement_settlement_mail'])
        existing_date = existing_claim.data.get("Settlement Mail to Accounts Date") if existing_claim else None
        if should_update_date('replacement_settlement_mail', existing_date):
            payload["Settlement Mail to Accounts Date"] = today_str

    if 'replacement_settled_accounts' in data: 
        is_settled = data['replacement_settled_accounts']
        payload["Settled With Accounts (Yes/No)"] = fmt_bool(is_settled)
        
        # BUSINESS RULE: If Settled with Accounts is checked, Settlement Mail to Accounts MUST be checked.
        if is_settled:
            payload["Settlement Mail to Accounts(Yes/No)"] = "Yes"
            
            # Only set the date if 'Settlement Mail to Accounts' was NOT ALREADY Yes
            existing_mail_checked = str(existing_claim.data.get("Settlement Mail to Accounts(Yes/No)")).strip().lower() if existing_claim else "no"
            if existing_mail_checked != "yes":
                payload["Settlement Mail to Accounts Date"] = today_str
    
    # Complete flag
    if 'complete' in data: payload["Complete (Yes/No)"] = fmt_bool(data['complete'])

    # CRITICAL: Mutual exclusivity of workflows
    # If status is 'Repair Completed', clear all Replacement Workflow data
    status_lower = (payload.get("Status") or "").strip().lower()
    
    if status_lower == "repair completed":
        payload["Customer Confirmation"] = ""
        payload["Approval Mail Received From Onsitego (Yes/No)"] = ""
        payload["Mail Sent To Store (Yes/No)"] = ""
        payload["Invoice Generated (Yes/No)"] = ""
        payload["Invoice Sent To Onsitego (Yes/No)"] = ""
        payload["Settlement Mail to Accounts(Yes/No)"] = ""
        payload["Settled With Accounts (Yes/No)"] = ""
    
    # If status is 'Replacement Approved', clear Repair Workflow data
    if "replacement" in status_lower and "approved" in status_lower:
        payload["Repair Feedback Completed (Yes/No)"] = ""

    # Sync
    try:
        sync_to_database_dict(payload, background=False)
    except Exception as e:
        print(f"Update Sync Error: {e}")
        return jsonify({"success": False})

    # --- GOOGLE SHEET WEBHOOK UPDATE ---
    web_app_url = os.environ.get("WEB_APP_URL")
    if web_app_url:
        import threading
        import requests
        
        def _update_sheet(url, update_payload, existing_claim_id, existing_sr):
            try:
                sheet_data = update_payload.copy()
                
                # Always include the Claim ID so Apps Script can find the row
                sheet_data["Claim ID"] = existing_claim_id
                
                # Remove empty SR No from payload — don't accidentally blank it in the sheet
                if "SR No" in sheet_data and not str(sheet_data.get("SR No", "")).strip():
                    del sheet_data["SR No"]
                
                # ALWAYS send legacy lookup hint if the old row had a CLM value in SR No.
                # This lets the Apps Script find the row via the SR No column even when
                # we're also writing a new real SR No at the same time.
                if existing_sr and existing_sr.startswith("CLM-"):
                    sheet_data["_legacy_sr_lookup"] = existing_sr
                
                real_sr = str(update_payload.get("SR No", "")).strip()
                print(f"[SHEET_UPDATE] Claim ID: {existing_claim_id} | SR No: {real_sr or '(unchanged)'} | Legacy lookup: {existing_sr or 'none'}")
                response = requests.post(url, json=sheet_data, timeout=15)
                print(f"[SHEET_UPDATE] Response: {response.status_code} - {response.text}")
            except Exception as e:
                print(f"[SHEET_UPDATE] Failed: {e}")
                
        # Include read-only fields so they populate in the Google Sheet when a claim is updated
        payload_for_sheet = payload.copy()
        if existing_claim:
            if existing_claim.customer_name: payload_for_sheet["Customer Name"] = existing_claim.customer_name
            if existing_claim.mobile_no: 
                payload_for_sheet["Mobile Number"] = existing_claim.mobile_no
                payload_for_sheet["Mobile"] = existing_claim.mobile_no
            if existing_claim.branch: payload_for_sheet["Branch"] = existing_claim.branch
            if existing_claim.product: payload_for_sheet["Product"] = existing_claim.product
            if existing_claim.issue: payload_for_sheet["Issue"] = existing_claim.issue
            
        # Get the existing SR No from before this update
        existing_sr_no = (existing_claim.sr_no or "").strip() if existing_claim else ""
        existing_claim_id = (existing_claim.claim_id or id) if existing_claim else id
        threading.Thread(target=_update_sheet, args=(web_app_url, payload_for_sheet, existing_claim_id, existing_sr_no)).start()

    # Invalidate Cache so next fetch gets fresh data
    global CLAIMS_CACHE
    CLAIMS_CACHE['last_updated'] = 0

    return jsonify({"success": True})

@app.route('/api/webhook/sheet-update', methods=['POST'])
def sheet_webhook_update():
    """
    Webhook endpoint to receive updates directly from Google Sheets via Apps Script onEdit trigger.
    """
    try:
        data = request.json
        if not data or not data.get("Claim ID"):
            return jsonify({"success": False, "message": "Missing Claim ID"}), 400
        
        # Sync the updated data from sheet into database
        sync_to_database_dict(data, background=False)
        
        # Invalidate Cache so next fetch gets fresh data
        global CLAIMS_CACHE
        CLAIMS_CACHE['last_updated'] = 0
        
        return jsonify({"success": True})
    except Exception as e:
        print(f"Webhook update error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

def sync_to_database_dict(payload, background=True):
    """
    Sends dict payload to PostgreSQL depending on background flag.
    Keys must match expected headers.
    """
    import datetime
    # Auto-add timestamp
    payload["Last Updated Timestamp"] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    def _sync():
        try:
            print(f"Starting Database Upsert for Claim {payload.get('Claim ID', 'Unknown')}...")
            from services.pg_sync import upsert_claim_to_postgres
            result = upsert_claim_to_postgres(payload)
            if result.get("success"):
                print(f"Upsert Status: Success")
            else:
                print(f"Database Upsert Failed: {result.get('error')}")
        except Exception as e:
            print(f"Database Upsert Failed Exception: {e}")

    # Start background thread
    if background:
        import threading
        threading.Thread(target=_sync).start()
    else:
        _sync()

@app.route('/api/notify-spare-parts/<string:id>', methods=['POST'])
@login_required
def notify_spare_parts(id):
    try:
        from services.whatsapp_service import send_whatsapp_message
        all_claims = fetch_claims_from_db()
        existing_claim = next((c for c in all_claims if str(c.claim_id) == str(id)), None)
        
        if not existing_claim:
            return jsonify({'success': False, 'message': 'Claim not found'})
            
        mobile = str(existing_claim.mobile_no) if existing_claim.mobile_no else ""
        c_name = str(existing_claim.customer_name).strip() if existing_claim.customer_name else "Customer"
        c_model = str(existing_claim.model).strip() if existing_claim.model else "your product"
        
        resp = send_whatsapp_message(
            mobile=mobile,
            template_name="myg_onsitego_part_order_main",
            params=[c_name, c_model]
        )
        
        if not resp.get("blocked") and resp.get("status_code") in [200, 201, 202]:
            timestamp = datetime.datetime.now().strftime('%d-%m-%Y %I:%M %p')
            note = f"[{timestamp}] Sent 'Spare Parts Pending' WhatsApp notification to customer."
            
            old_history = existing_claim.follow_up_notes or ""
            new_history = f"{old_history}\n{note}" if str(old_history).strip() else note
            
            payload = {
                "Claim ID": id,
                "Follow Up - Notes": new_history,
            }
            if PG_AVAILABLE:
                from services.pg_sync import upsert_claim_to_postgres
                upsert_claim_to_postgres(payload)
                invalidate_cache()
                
            return jsonify({'success': True, 'message': 'Notification sent successfully!'})
        else:
            return jsonify({'success': False, 'message': resp.get("error", "Failed to send notification")})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# ----------------------
# DEBUG ENDPOINT
# ----------------------
@app.route('/debug/sheet-columns')
@admin_required
def debug_sheet_columns():
    """Debug endpoint to see actual column names and sample data"""
    try:
        claims = fetch_claims_from_db()
        if len(claims) > 0:
            first_claim = claims[0]
            return jsonify({
                'success': True,
                'sample_claim_id': first_claim.claim_id,
                'all_columns': list(first_claim.data.keys()),
                'replacement_columns': {
                    'Replacement: Confirmation Pending': first_claim.data.get('Replacement: Confirmation Pending'),
                    'Replacement: OSG Approval': first_claim.data.get('Replacement: OSG Approval'),
                    'Replacement: Mail to Store': first_claim.data.get('Replacement: Mail to Store'),
                    'Replacement: Invoice Generated': first_claim.data.get('Replacement: Invoice Generated'),
                    'Replacement: Invoice Sent to OSG': first_claim.data.get('Replacement: Invoice Sent to OSG'),
                    'Replacement: Settled with Accounts': first_claim.data.get('Replacement: Settled with Accounts'),
                    'Complete': first_claim.data.get('Complete')
                }
            })
        return jsonify({'success': False, 'message': 'No claims found'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/debug/gst-claims')
@admin_required
def debug_gst_claims():
    """Debug endpoint to see all GST Invoice Billed claims and their date fields"""
    try:
        claims = fetch_claims_from_db(force_refresh=True)
        now = get_ist_now().replace(tzinfo=None)
        gst_claims = []
        for c in claims:
            if c.invoice_generated:
                status = (c.status or '').strip().lower()
                if 'replacement' in status or c.mail_sent_to_store:
                    if not c.invoice_sent_osg and not c.settlement_mail_accounts and not c.settled_with_accounts:
                        age = (now - c.created_at.replace(tzinfo=None)).days if c.created_at else 0
                        gst_claims.append({
                            'claim_id': c.claim_id,
                            'status': c.status,
                            'created_at': str(c.created_at)[:10],
                            'claim_settled_date': c.claim_settled_date,
                            'invoice_generated_date': c.invoice_generated_date,
                            'mail_sent_to_store_date': c.mail_sent_to_store_date,
                            'raw_invoice_generated_date': c.data.get('Invoice Generated Date'),
                            'raw_mail_sent_to_store_date': c.data.get('Mail Sent To Store Date'),
                            'submission_age_days': age,
                            'all_date_columns': {k: v for k, v in c.data.items() if 'date' in k.lower() or 'Date' in k}
                        })
        return jsonify({'success': True, 'count': len(gst_claims), 'claims': gst_claims})
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e), 'trace': traceback.format_exc()})

# ----------------------
# ANALYTICS ROUTES
# ----------------------
@app.route('/analytics')
@admin_required
def analytics_dashboard():
    return render_template('analytics.html')

# ----------------------
# CLAIM STATUS (Customer Care)
# ----------------------
@app.route('/claim-status')
@login_required
def claim_status():
    claims = fetch_claims_from_db()
    now = get_ist_now().replace(tzinfo=None)

    # KPI stats
    total = len(claims)
    pending = len([c for c in claims if not c.complete])
    completed = len([c for c in claims if c.complete])
    tat_values = [c.tat for c in claims if c.tat is not None and isinstance(c.tat, int)]
    avg_tat = round(sum(tat_values) / len(tat_values)) if tat_values else 0

    report_stats = {
        'pending': {'lt5': 0, 'gt5': 0, 'gt10': 0, 'total': 0},
        'completed': 0,
        'rejected': 0,
        'replacement_mail': {'lt5': 0, 'gt5': 0, 'gt10': 0, 'total': 0},
        'gst_invoice': {'lt5': 0, 'gt5': 0, 'gt10': 0, 'total': 0},
        'pending_settlement_osg': {'lt5': 0, 'gt5': 0, 'gt10': 0, 'total': 0},
        'settlement_mail_accounts': {'lt5': 0, 'gt5': 0, 'gt10': 0, 'total': 0},
        'settled_accounts': 0,
        'grand_total_status': 0,
        'grand_total_replacement': 0,
        'report_date': now.strftime('%d-%m-%Y')
    }

    for c in claims:
        age = (now - c.created_at.replace(tzinfo=None)).days if c.created_at else 0
        settled_date_raw = c.claim_settled_date
        repl_age = age
        if settled_date_raw and str(settled_date_raw).strip() not in ('', 'nan', 'None'):
            try:
                settled_dt = datetime.datetime.strptime(str(settled_date_raw).strip()[:10], '%Y-%m-%d')
                repl_age = (now - settled_dt).days
            except Exception:
                try:
                    settled_dt = datetime.datetime.strptime(str(settled_date_raw).strip()[:10], '%d-%m-%Y')
                    repl_age = (now - settled_dt).days
                except Exception:
                    repl_age = age

        status = (c.status or "").strip().lower()

        if status == "rejected":
            report_stats['rejected'] += 1
            report_stats['grand_total_status'] += 1
        elif c.complete or status in ["repair completed", "closed", "no issue/oncall resolution",
                                       "no issue", "oncall resolution", "cancelled"]:
            report_stats['completed'] += 1
            report_stats['grand_total_status'] += 1
        else:
            report_stats['pending']['total'] += 1
            report_stats['grand_total_status'] += 1
            if age <= 5:
                report_stats['pending']['lt5'] += 1
            elif age <= 10:
                report_stats['pending']['gt5'] += 1
            else:
                report_stats['pending']['gt10'] += 1

        if "replacement" in status or c.mail_sent_to_store:
            if c.settled_with_accounts:
                report_stats['settled_accounts'] += 1
                report_stats['grand_total_replacement'] += 1
            elif c.settlement_mail_accounts:
                report_stats['settlement_mail_accounts']['total'] += 1
                report_stats['grand_total_replacement'] += 1
                if repl_age <= 5: report_stats['settlement_mail_accounts']['lt5'] += 1
                elif repl_age <= 10: report_stats['settlement_mail_accounts']['gt5'] += 1
                else: report_stats['settlement_mail_accounts']['gt10'] += 1
            elif c.invoice_sent_osg:
                report_stats['pending_settlement_osg']['total'] += 1
                report_stats['grand_total_replacement'] += 1
                if repl_age <= 5: report_stats['pending_settlement_osg']['lt5'] += 1
                elif repl_age <= 10: report_stats['pending_settlement_osg']['gt5'] += 1
                else: report_stats['pending_settlement_osg']['gt10'] += 1
            elif c.invoice_generated:
                report_stats['gst_invoice']['total'] += 1
                report_stats['grand_total_replacement'] += 1
                # Age from Invoice Generated Date (when GST invoice was billed)
                # Fallback chain: Invoice Generated Date → Mail Sent To Store Date → repl_age
                def _parse_date(raw):
                    if not raw or str(raw).strip() in ('', 'nan', 'None'): return None
                    s = str(raw).strip()[:10]
                    dt = None
                    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
                        try:
                            dt = datetime.datetime.strptime(s, fmt)
                            break
                        except (ValueError, TypeError): continue
                    
                    # If it's a future date, Google Sheets likely interpreted DD-MM as MM-DD
                    if dt and (dt - now).days > 1:
                        try:
                            if '-' in s and len(s.split('-')[0]) == 4:
                                dt = datetime.datetime.strptime(s, '%Y-%d-%m')
                        except (ValueError, TypeError): pass
                    return dt

                inv_gen_dt = _parse_date(c.invoice_generated_date)
                if inv_gen_dt:
                    gst_age = max(0, (now - inv_gen_dt).days)
                else:
                    store_dt = _parse_date(c.mail_sent_to_store_date)
                    if store_dt:
                        gst_age = max(0, (now - store_dt).days)
                    else:
                        gst_age = max(0, repl_age)

                if gst_age <= 5: report_stats['gst_invoice']['lt5'] += 1
                elif gst_age <= 10: report_stats['gst_invoice']['gt5'] += 1
                else: report_stats['gst_invoice']['gt10'] += 1
            else:
                report_stats['replacement_mail']['total'] += 1
                report_stats['grand_total_replacement'] += 1
                if repl_age <= 5:
                    report_stats['replacement_mail']['lt5'] += 1
                elif repl_age <= 10:
                    report_stats['replacement_mail']['gt5'] += 1
                else:
                    report_stats['replacement_mail']['gt10'] += 1

    return render_template('claim_status.html', report_stats=report_stats,
                           total=total, pending=pending, completed=completed, avg_tat=avg_tat)

@app.route('/api/claim-status-lookup', methods=['POST'])
@login_required
def claim_status_lookup():
    """Search claims by mobile number or claim ID for customer care"""
    try:
        data = request.json
        if not data:
            return jsonify({'success': False, 'message': 'Invalid request body.'})

        # --- INPUT VALIDATION ---
        try:
            search_type = validate_search_type(data.get('search_type', 'mobile'))
            search_value = sanitize_text(data.get('search_value', ''), max_len=50,
                                         field_name='search_value')
        except ValueError as e:
            return jsonify({'success': False, 'message': str(e)})

        if not search_value:
            return jsonify({'success': False, 'message': 'Search value is required'})

        # Additional type-specific validation
        if search_type == 'mobile':
            try:
                search_value = validate_mobile(search_value)
            except ValueError as e:
                return jsonify({'success': False, 'message': str(e)})
        # --- END VALIDATION ---
        
        claims = fetch_claims_from_db()
        matched = []
        
        def parse_bool(val):
            if val is None or val == '':
                return False
            return str(val).strip().lower() in ['yes', 'true', '1']
        
        for c in claims:
            if search_type == 'mobile':
                mobile = str(c.mobile_no or '').strip()
                if mobile == search_value:
                    matched.append(c)
            else:
                claim_id = str(c.claim_id or '').strip().lower()
                if search_value.lower() in claim_id:
                    matched.append(c)
        
        if not matched:
            return jsonify({'success': False, 'message': 'No claims found'})
        
        results = []
        for c in matched:
            results.append({
                'claim_id': c.claim_id,
                'submitted_date': c.created_at.strftime('%Y-%m-%d') if c.created_at else '',
                'customer_name': c.customer_name or '',
                'mobile_number': c.mobile_no or '',
                'product': c.model or '',
                'model': c.model or '',
                'status': c.status or '',
                'osid': c.osid or '',
                'sr_no': c.sr_no or '',
                'invoice_no': c.invoice_no or '',
                'branch': c.branch or '',
                'issue': c.issue or '',
                'claim_settled_date': c.claim_settled_date or '',
                'follow_up_notes': c.follow_up_notes or '',
                'follow_up_date': c.follow_up_date or '',
                'tat': c.tat,
                'complete': c.complete,
                'replacement_confirmation': parse_bool(c.data.get("Customer Confirmation")),
                'replacement_osg_approval': parse_bool(c.data.get("Approval Mail Received From Onsitego (Yes/No)")),
                'replacement_mail_store': parse_bool(c.data.get("Mail Sent To Store (Yes/No)")),
                'replacement_invoice_gen': parse_bool(c.data.get("Invoice Generated (Yes/No)")),
                'replacement_invoice_sent': parse_bool(c.data.get("Invoice Sent To Onsitego (Yes/No)")),
                'settlement_mail_accounts': parse_bool(c.data.get("Settlement Mail to Accounts(Yes/No)")),
                'replacement_settled_accounts': parse_bool(c.data.get("Settled With Accounts (Yes/No)")),
            })
        
        return jsonify({'success': True, 'claims': results})
    except Exception as e:
        print(f"Claim status lookup error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/analytics-data')
@login_required
def get_analytics_data():
    """
    Fetch and transform claims data for analytics dashboard
    Returns structured JSON with all necessary fields
    """
    try:
        claims = fetch_claims_from_db()
        
        # Transform claims for analytics
        analytics_claims = []
        for claim in claims:
            # Calculate TAT if settled
            tat = None
            if claim.claim_settled_date and (claim.data.get("Date") or claim.data.get("date") or claim.data.get("Submitted Date")):
                try:
                    s_date = claim.data.get("Date") or claim.data.get("date") or claim.data.get("Submitted Date")
                    submitted = datetime.datetime.strptime(str(s_date).split()[0], '%Y-%m-%d')
                    settled = datetime.datetime.strptime(str(claim.claim_settled_date).split()[0], '%Y-%m-%d')
                    tat = (settled - submitted).days
                except:
                    tat = None
            
            # Get replacement workflow fields
            def parse_bool(val):
                if val is None or val == '':
                    return False
                return str(val).strip().lower() in ['yes', 'true', '1']
            
            
            # Format mobile number to ensure it's a clean string
            mobile_raw = claim.mobile_no or ''
            if mobile_raw:
                # Convert to string and remove decimal points (e.g., "8589852744.0" -> "8589852744")
                mobile_str = str(mobile_raw).strip()
                if '.' in mobile_str:
                    mobile_str = mobile_str.split('.')[0]
                mobile_formatted = mobile_str
            else:
                mobile_formatted = ''
            
            s_date_final = claim.data.get("Date") or claim.data.get("date") or claim.data.get("Submitted Date", '')
            analytics_claims.append({
                'claim_id': claim.claim_id or '',
                'submitted_date': str(s_date_final).split()[0] if s_date_final else '',
                'customer_name': claim.customer_name or '',
                'mobile_number': mobile_formatted,
                'address': claim.address or '',
                'branch': claim.data.get("Branch") or claim.data.get("Branch Name") or 'Main Branch',
                'product': claim.data.get("Product", claim.model) or '',
                'model': claim.model or '',
                'invoice_number': claim.invoice_no or '',
                'serial_number': claim.serial_no or '',
                'sr_no': claim.sr_no or '',
                'osid': claim.osid or '',
                'issue': claim.issue or '',
                'status': claim.status or 'Unknown',
                'remarks': claim.remarks or '',
                'follow_up_notes': claim.follow_up_notes or '',
                'claim_settled_date': claim.claim_settled_date or '',
                'tat': tat,
                
                # Replacement workflow fields (Columns O-T) - Use actual sheet column names
                'replacement_confirmation': parse_bool(claim.data.get("Customer Confirmation")),
                'replacement_osg_approval': parse_bool(claim.data.get("Approval Mail Received From Onsitego (Yes/No)")),
                'replacement_mail_store': parse_bool(claim.data.get("Mail Sent To Store (Yes/No)")),
                'replacement_invoice_gen': parse_bool(claim.data.get("Invoice Generated (Yes/No)")),
                'replacement_invoice_sent': parse_bool(claim.data.get("Invoice Sent To Onsitego (Yes/No)")),
                'settlement_mail_accounts': parse_bool(claim.data.get("Settlement Mail to Accounts(Yes/No)")),
                'replacement_settled_accounts': parse_bool(claim.data.get("Settled With Accounts (Yes/No)")),
                
                # Complete flag
                'complete': claim.complete
            })
        
        return jsonify({
            'success': True,
            'claims': analytics_claims,
            'total': len(analytics_claims)
        })
        
    except Exception as e:
        print(f"Analytics data error: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'claims': []
        })

# ----------------------
# REPORTS & TOOLS ROUTES
# ----------------------
@app.route('/reports')
@admin_required
def reports_tools():
    today = datetime.datetime.now().strftime('%Y-%m-%d')
    month_start = datetime.datetime.now().replace(day=1).strftime('%Y-%m-%d')
    return render_template('reports_tools.html', today=today, month_start=month_start)

@app.route('/reports/generate_1', methods=['POST'])
@login_required
def generate_report_1():
    import xlsxwriter
    try:
        report_date_str = request.form.get('report_date', '').strip()
        prev_date_str   = request.form.get('prev_date', '').strip()

        # --- INPUT VALIDATION ---
        try:
            report_date_str = validate_date_str(report_date_str, field_name='report_date')
            prev_date_str   = validate_date_str(prev_date_str,   field_name='prev_date')
            if not report_date_str or not prev_date_str:
                flash("Report date and previous date are required.", "error")
                return redirect(url_for('reports_tools'))
        except ValueError as e:
            flash(str(e), "error")
            return redirect(url_for('reports_tools'))

        curr_sales_file    = request.files.get('curr_sales')
        prev_sales_file    = request.files.get('prev_sales')
        product_sales_file = request.files.get('product_sales')

        if not curr_sales_file or not product_sales_file or not prev_sales_file:
            flash("All files (Current Sales, Previous Sales, Product Sales) are required.", "error")
            return redirect(url_for('reports_tools'))

        for upload, label in [(curr_sales_file, 'Current Sales'), (prev_sales_file, 'Previous Sales'), (product_sales_file, 'Product Sales')]:
            try:
                validate_file_upload(upload, field_name=label)
            except ValueError as e:
                flash(str(e), "error")
                return redirect(url_for('reports_tools'))
        # --- END VALIDATION ---

        from services.pg_sync import _get_connection
        import pandas as pd
        
        # Load Defaults from DB with Fallback to Excel
        store_list_path = os.path.join(BASE_DIR, "myG All Store.xlsx")
        rbm_path = os.path.join(BASE_DIR, "RBM,BDM,BRANCH.xlsx")

        try:
            conn = _get_connection()
            future_store_df = pd.read_sql('SELECT store as "Store" FROM myg_all_store', conn)
            if future_store_df.empty: raise Exception("Empty DB table myg_all_store")
        except Exception as e:
            if not os.path.exists(store_list_path):
                flash("Default Store file not found on server or DB.", "error")
                return redirect(url_for('reports_tools'))
            future_store_df = timed_excel_read(store_list_path, "Store List", engine='openpyxl')
            if 'Store' not in future_store_df.columns:
                if 'Store Name' in future_store_df.columns: future_store_df.rename(columns={'Store Name': 'Store'}, inplace=True)
                elif 'Branch' in future_store_df.columns: future_store_df.rename(columns={'Branch': 'Store'}, inplace=True)

        try:
            conn = _get_connection()
            rbm_df = pd.read_sql('SELECT branch as "Store", rbm as "RBM", bdm as "BDM" FROM rbm_bdm_branch', conn)
            if rbm_df.empty: raise Exception("Empty DB table rbm_bdm_branch")
        except Exception as e:
            if not os.path.exists(rbm_path):
                flash("Default RBM file not found on server or DB.", "error")
                return redirect(url_for('reports_tools'))
            rbm_df = timed_excel_read(rbm_path, "RBM Mapping", engine='openpyxl')
            if 'Store' not in rbm_df.columns:
                if 'Store Name' in rbm_df.columns: rbm_df.rename(columns={'Store Name': 'Store'}, inplace=True)
                elif 'Branch' in rbm_df.columns: rbm_df.rename(columns={'Branch': 'Store'}, inplace=True)

        # Process logic from snippet
        # Optimization: Read only needed columns
        # Needed: Branch (->Store), DATE, QUANTITY, AMOUNT
        try:
             book1_df = timed_excel_read(curr_sales_file, "Current Sales", engine='openpyxl')
        except Exception:
             curr_sales_file.seek(0)
             book1_df = timed_excel_read(curr_sales_file, "Current Sales (Fallback)", engine='openpyxl')

        book1_df = book1_df.loc[:, ~book1_df.columns.duplicated()]
        book1_df.rename(columns={'Branch': 'Store', 'Date': 'DATE'}, inplace=True)
        book1_df['DATE'] = pd.to_datetime(book1_df['DATE'], dayfirst=True, errors='coerce')
        book1_df = book1_df.dropna(subset=['DATE'])
        rbm_df.rename(columns={'Branch': 'Store'}, inplace=True)

        try:
            product_df = timed_excel_read(product_sales_file, "Product Sales", engine='openpyxl')
        except:
            product_sales_file.seek(0)
            product_df = timed_excel_read(product_sales_file, "Product Sales (Fallback)", engine='openpyxl')
            
        product_df = product_df.loc[:, ~product_df.columns.duplicated()]
        product_df.rename(columns={'Branch': 'Store', 'Date': 'DATE', 'Sold Price': 'AMOUNT'}, inplace=True)
        product_df['DATE'] = pd.to_datetime(product_df['DATE'], dayfirst=True, errors='coerce')
        product_df = product_df.dropna(subset=['DATE'])
        if 'QUANTITY' not in product_df.columns:
            product_df['QUANTITY'] = 1

        today = pd.to_datetime(report_date_str)
        mtd_df = book1_df[book1_df['DATE'].dt.month == today.month]
        today_df = mtd_df[mtd_df['DATE'].dt.date == today.date()]
        
        today_agg = today_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(columns={'QUANTITY': 'FTD Count', 'AMOUNT': 'FTD Value'})
        mtd_agg = mtd_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(columns={'QUANTITY': 'MTD Count', 'AMOUNT': 'MTD Value'})

        product_mtd_df = product_df[product_df['DATE'].dt.month == today.month]
        product_today_df = product_mtd_df[product_mtd_df['DATE'].dt.date == today.date()]
        product_today_agg = product_today_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(columns={'QUANTITY': 'Product_FTD_Count', 'AMOUNT': 'Product_FTD_Amount'})
        product_mtd_agg = product_mtd_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(columns={'QUANTITY': 'Product_MTD_Count', 'AMOUNT': 'Product_MTD_Amount'})

        try:
            prev_df = timed_excel_read(prev_sales_file, "Previous Sales", engine='openpyxl')
        except:
             prev_sales_file.seek(0)
             prev_df = timed_excel_read(prev_sales_file, "Previous Sales (Fallback)", engine='openpyxl')

        prev_df = prev_df.loc[:, ~prev_df.columns.duplicated()]
        prev_df.rename(columns={'Branch': 'Store', 'Date': 'DATE'}, inplace=True)
        prev_df['DATE'] = pd.to_datetime(prev_df['DATE'], dayfirst=True, errors='coerce')
        prev_df = prev_df.dropna(subset=['DATE'])
        prev_month = pd.to_datetime(prev_date_str)
        prev_mtd_df = prev_df[prev_df['DATE'].dt.month == prev_month.month]
        prev_mtd_agg = prev_mtd_df.groupby('Store', as_index=False).agg({'AMOUNT': 'sum'}).rename(columns={'AMOUNT': 'PREV MONTH SALE'})

        # Extract Store arrays, dropping NAs, and flatten to 1D to prevent "Grouper not 1-dimensional" error
        fs_stores = future_store_df['Store'].dropna().astype(str).values if 'Store' in future_store_df.columns else []
        b1_stores = book1_df['Store'].dropna().astype(str).values if 'Store' in book1_df.columns else []
        p_stores = product_df['Store'].dropna().astype(str).values if 'Store' in product_df.columns else []
        
        import numpy as np
        all_unique_stores = np.unique(np.concatenate([fs_stores, b1_stores, p_stores]))
        all_stores = pd.DataFrame({'Store': all_unique_stores})
        report_df = all_stores.merge(today_agg, on='Store', how='left') \
                                .merge(mtd_agg, on='Store', how='left') \
                                .merge(product_today_agg, on='Store', how='left') \
                                .merge(product_mtd_agg, on='Store', how='left') \
                                .merge(prev_mtd_agg, on='Store', how='left') \
                                .merge(rbm_df[['Store', 'RBM']], on='Store', how='left')

        required_columns = ['Store', 'FTD Count', 'FTD Value', 'Product_FTD_Amount', 'MTD Count', 'MTD Value', 'Product_MTD_Amount', 'PREV MONTH SALE', 'RBM']
        for col in required_columns:
            if col not in report_df.columns:
                report_df[col] = 0
        report_df = report_df.rename(columns={'Store': 'Store Name'})

        cols_to_fill = ['FTD Count', 'FTD Value', 'MTD Count', 'MTD Value', 'Product_FTD_Count', 'Product_FTD_Amount', 'Product_MTD_Count', 'Product_MTD_Amount', 'PREV MONTH SALE']
        # Only fill available columns
        available_fill = [c for c in cols_to_fill if c in report_df.columns]
        report_df[available_fill] = report_df[available_fill].fillna(0).astype(int)

        report_df['DIFF %'] = report_df.apply(
            lambda x: round(((x['MTD Value'] - x['PREV MONTH SALE']) / x['PREV MONTH SALE']) * 100, 2) if x['PREV MONTH SALE'] != 0 else 0,
            axis=1
        )
        report_df['ASP'] = report_df.apply(
            lambda x: round(x['MTD Value'] / x['MTD Count'], 2) if x['MTD Count'] != 0 else 0,
            axis=1
        )
        report_df['FTD Value Conversion'] = report_df.apply(
            lambda x: round((x['FTD Value'] / x['Product_FTD_Amount']) * 100, 2) if x['Product_FTD_Amount'] != 0 else 0,
            axis=1
        )
        report_df['MTD Value Conversion'] = report_df.apply(
            lambda x: round((x['MTD Value'] / x['Product_MTD_Amount']) * 100, 2) if x['Product_MTD_Amount'] != 0 else 0,
            axis=1
        )

        # Excel Generation with Complete Streamlit Formatting
        excel_output = io.BytesIO()
        with pd.ExcelWriter(excel_output, engine='xlsxwriter') as writer:
            workbook = writer.book

            colors_palette = {
                'primary_blue': '#1E3A8A',
                'light_blue': '#DBEAFE',
                'success_green': '#065F46',
                'light_green': '#D1FAE5',
                'warning_orange': '#EA580C',
                'light_orange': '#FED7AA',
                'danger_red': '#DC2626',
                'light_red': '#FEE2E2',
                'accent_purple': '#7C3AED',
                'light_purple': '#EDE9FE',
                'neutral_gray': '#6B7280',
                'light_gray': '#F9FAFB',
                'white': '#FFFFFF',
                'dark_blue': '#0F172A',
                'mint_green': '#10B981',
                'light_mint': '#ECFDF5',
                'royal_blue': '#3B82F6',
                'light_royal': '#EBF8FF'
            }

            formats = {
                'title': workbook.add_format({
                    'bold': True, 'font_size': 16, 'font_color': colors_palette['primary_blue'],
                    'align': 'center', 'valign': 'vcenter', 'bg_color': colors_palette['white'],
                    'border': 1, 'border_color': colors_palette['primary_blue']
                }),
                'subtitle': workbook.add_format({
                    'bold': True, 'font_size': 12, 'font_color': colors_palette['neutral_gray'],
                    'align': 'center', 'valign': 'vcenter', 'italic': True
                }),
                'header_main': workbook.add_format({
                    'bold': True, 'font_size': 11, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['primary_blue'], 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['primary_blue'], 'text_wrap': True
                }),
                'header_secondary': workbook.add_format({
                    'bold': True, 'font_size': 10, 'font_color': colors_palette['primary_blue'],
                    'bg_color': colors_palette['light_blue'], 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['primary_blue']
                }),
                'data_normal': workbook.add_format({
                    'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['white']
                }),
                'data_alternate': workbook.add_format({
                    'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_gray']
                }),
                'data_store_name': workbook.add_format({
                    'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['white'], 'indent': 1
                }),
                'data_store_name_alt': workbook.add_format({
                    'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_gray'], 'indent': 1
                }),
                'conversion_low': workbook.add_format({
                    'font_size': 10, 'font_color': colors_palette['danger_red'], 'bg_color': colors_palette['light_red'],
                    'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['danger_red'], 'num_format': '0.00%', 'bold': True
                }),
                'conversion_green': workbook.add_format({
                    'bold': True, 'font_size': 10, 'font_color': colors_palette['success_green'],
                    'bg_color': colors_palette['light_green'], 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['success_green'], 'num_format': '0.00%'
                }),
                'conversion_format': workbook.add_format({
                    'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'num_format': '0.00%'
                }),
                'conversion_format_alt': workbook.add_format({
                    'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal'], 'num_format': '0.00%'
                }),
                'total_row': workbook.add_format({
                    'bold': True, 'font_size': 11, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['mint_green'], 'align': 'center', 'valign': 'vcenter',
                    'border': 2, 'border_color': colors_palette['mint_green']
                }),
                'total_label': workbook.add_format({
                    'bold': True, 'font_size': 11, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['mint_green'], 'align': 'center', 'valign': 'vcenter',
                    'border': 2, 'border_color': colors_palette['mint_green']
                }),
                'rbm_title': workbook.add_format({
                    'bold': True, 'font_size': 18, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['dark_blue'], 'align': 'center', 'valign': 'vcenter',
                    'border': 2, 'border_color': colors_palette['dark_blue']
                }),
                'rbm_subtitle': workbook.add_format({
                    'bold': True, 'font_size': 11, 'font_color': colors_palette['dark_blue'],
                    'bg_color': colors_palette['light_royal'], 'align': 'center', 'valign':' vcenter',
                    'border': 1, 'border_color': colors_palette['royal_blue'], 'italic': True
                }),
                'rbm_header': workbook.add_format({
                    'bold': True, 'font_size': 11, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['royal_blue'], 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['royal_blue'], 'text_wrap': True
                }),
                'rbm_data_normal': workbook.add_format({
                    'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['white']
                }),
                'rbm_data_alternate': workbook.add_format({
                    'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal']
                }),
                'rbm_store_name': workbook.add_format({
                    'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['white'], 'indent': 1
                }),
                'rbm_store_name_alt': workbook.add_format({
                    'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal'], 'indent': 1
                }),
                'rbm_conversion_low': workbook.add_format({
                    'font_size': 10, 'font_color': colors_palette['danger_red'], 'bg_color': colors_palette['light_red'],
                    'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['danger_red'], 'num_format': '0.00%', 'bold': True
                }),
                'rbm_conversion_green': workbook.add_format({
                    'bold': True, 'font_size': 10, 'font_color': colors_palette['success_green'],
                    'bg_color': colors_palette['light_green'], 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['success_green'], 'num_format': '0.00%'
                }),
                'rbm_conversion_format': workbook.add_format({
                    'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'num_format': '0.00%'
                }),
                'rbm_conversion_format_alt': workbook.add_format({
                    'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal'], 'num_format': '0.00%'
                }),
                'rbm_total': workbook.add_format({
                    'bold': True, 'font_size': 12, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['mint_green'], 'align': 'center', 'valign': 'vcenter',
                    'border': 2, 'border_color': colors_palette['mint_green']
                }),
                'rbm_total_label': workbook.add_format({
                    'bold': True, 'font_size': 12, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['mint_green'], 'align': 'center', 'valign': 'vcenter',
                    'border': 2, 'border_color': colors_palette['mint_green']
                }),
                'rbm_summary': workbook.add_format({
                    'bold': True, 'font_size': 10, 'font_color': colors_palette['royal_blue'],
                    'bg_color': colors_palette['light_royal'], 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['royal_blue']
                }),
                'rbm_performance': workbook.add_format({
                    'bold': True, 'font_size': 10, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['accent_purple'], 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['accent_purple']
                }),
                'asp_format': workbook.add_format({
                    'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'num_format': '₹#,##0.00'
                }),
                'asp_format_alt': workbook.add_format({
                    'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal'], 'num_format': '₹#,##0.00'
                }),
                'asp_total': workbook.add_format({
                    'bold': True, 'font_size': 12, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['mint_green'], 'align': 'center', 'valign': 'vcenter',
                    'border': 2, 'border_color': colors_palette['mint_green'], 'num_format': '₹#,##0.00'
                })
            }

            # Set IST timezone
            ist_time = datetime.datetime.now(pytz.timezone('Asia/Kolkata'))

            # ALL STORES SHEET
            all_data = report_df.sort_values('MTD Value', ascending=False)
            worksheet = workbook.add_worksheet("All Stores")

            # Headers
            headers = ['Store Name', 'FTD Count', 'FTD Value', 'FTD Value Conversion', 'MTD Count', 'MTD Value', 'MTD Value Conversion', 'PREV MONTH SALE', 'DIFF %', 'ASP']
            
            # Optimized: Use fixed column widths for speed (works for most data)
            column_widths = [25, 12, 12, 20, 12, 12, 20, 18, 10, 15]
            for i, width in enumerate(column_widths):
                worksheet.set_column(i, i, width)

            # Title and subtitle
            worksheet.merge_range(0, 0, 0, len(headers) - 1, "OSG All Stores Sales Report", formats['title'])
            worksheet.merge_range(1, 0, 1, len(headers) - 1, f"Report Generated: {ist_time.strftime('%d %B %Y %I:%M %p IST')}", formats['subtitle'])

            # Summary statistics
            total_stores = len(all_data)
            active_stores = len(all_data[all_data['FTD Count'] > 0])
            inactive_stores = total_stores - active_stores
            worksheet.merge_range(3, 0, 3, 1, "📊 SUMMARY", formats['header_secondary'])
            worksheet.merge_range(3, 2, 3, len(headers) - 1, f"Total: {total_stores} | Active: {active_stores} | Inactive: {inactive_stores}", formats['data_normal'])

            # Column headers
            for col, header in enumerate(headers):
                worksheet.write(5, col, header, formats['header_main'])

            # Cache formats for speed (avoid dictionary lookups in loop)
            fmt_data_normal = formats['data_normal']
            fmt_data_alternate = formats['data_alternate']
            fmt_store_normal = formats['data_store_name']
            fmt_store_alt = formats['data_store_name_alt']
            fmt_conv_normal = formats['conversion_format']
            fmt_conv_alt = formats['conversion_format_alt']
            fmt_conv_green = formats['conversion_green']
            fmt_conv_low = formats['conversion_low']
            fmt_asp_normal = formats['asp_format']
            fmt_asp_alt = formats['asp_format_alt']

            # Data rows with alternating colors (OPTIMIZED)
            for row_idx, (_, row) in enumerate(all_data.iterrows(), start=6):
                is_alternate = (row_idx - 6) % 2 == 1
                data_format = fmt_data_alternate if is_alternate else fmt_data_normal
                store_format = fmt_store_alt if is_alternate else fmt_store_normal
                asp_format = fmt_asp_alt if is_alternate else fmt_asp_normal
                conversion_format = fmt_conv_alt if is_alternate else fmt_conv_normal
                
                # Prepare row data - write entire row at once for speed
                row_data = [
                    row['Store Name'],
                    int(row['FTD Count']),
                    int(row['FTD Value']),
                    row['FTD Value Conversion'] / 100,  # Will format separately
                    int(row['MTD Count']),
                    int(row['MTD Value']),
                    row['MTD Value Conversion'] / 100,  # Will format separately
                    int(row['PREV MONTH SALE']),
                    f"{row['DIFF %']}%",
                    row['ASP']
                ]
                
                # Write entire row with default format first
                worksheet.write_row(row_idx, 0, row_data, data_format)
                
                # Override specific cells with special formats
                worksheet.write(row_idx, 0, row['Store Name'], store_format)
                
                # FTD Conversion - conditional formatting
                ftd_conversion = row['FTD Value Conversion']
                if ftd_conversion > 2:
                    worksheet.write(row_idx, 3, ftd_conversion / 100, fmt_conv_green)
                elif ftd_conversion < 2:
                    worksheet.write(row_idx, 3, ftd_conversion / 100, fmt_conv_low)
                else:
                    worksheet.write(row_idx, 3, ftd_conversion / 100, conversion_format)
                
                # MTD Conversion - conditional formatting
                mtd_conversion = row['MTD Value Conversion']
                if mtd_conversion > 2:
                    worksheet.write(row_idx, 6, mtd_conversion / 100, fmt_conv_green)
                elif mtd_conversion < 2:
                    worksheet.write(row_idx, 6, mtd_conversion / 100, fmt_conv_low)
                else:
                    worksheet.write(row_idx, 6, mtd_conversion / 100, conversion_format)
                
                # ASP with currency format
                worksheet.write(row_idx, 9, row['ASP'], asp_format)

            # Total row
            total_row = len(all_data) + 7
            worksheet.write(total_row, 0, '🎯 TOTAL', formats['total_label'])
            worksheet.write(total_row, 1, all_data['FTD Count'].sum(), formats['total_row'])
            worksheet.write(total_row, 2, all_data['FTD Value'].sum(), formats['total_row'])
            total_ftd_conversion = round((all_data['FTD Value'].sum() / all_data['Product_FTD_Amount'].sum()) * 100, 2) if all_data['Product_FTD_Amount'].sum() != 0 else 0
            worksheet.write(total_row, 3, f"{total_ftd_conversion}%", formats['total_row'])
            worksheet.write(total_row, 4, all_data['MTD Count'].sum(), formats['total_row'])
            worksheet.write(total_row, 5, all_data['MTD Value'].sum(), formats['total_row'])
            total_mtd_conversion = round((all_data['MTD Value'].sum() / all_data['Product_MTD_Amount'].sum()) * 100, 2) if all_data['Product_MTD_Amount'].sum() != 0 else 0
            worksheet.write(total_row, 6, f"{total_mtd_conversion}%", formats['total_row'])
            worksheet.write(total_row, 7, all_data['PREV MONTH SALE'].sum(), formats['total_row'])
            total_diff = round(((all_data['MTD Value'].sum() - all_data['PREV MONTH SALE'].sum()) / all_data['PREV MONTH SALE'].sum()) * 100, 2) if all_data['PREV MONTH SALE'].sum() != 0 else 0
            worksheet.write(total_row, 8, f"{total_diff}%", formats['total_row'])
            total_asp = round(all_data['MTD Value'].sum() / all_data['MTD Count'].sum(), 2) if all_data['MTD Count'].sum() != 0 else 0
            worksheet.write(total_row, 9, total_asp, formats['asp_total'])

            # Top performer insight
            if len(all_data) > 0:
                top_performer = all_data.iloc[0]
                insights_row = total_row + 2
                worksheet.merge_range(insights_row, 0, insights_row, len(headers) - 1,
                                    f"🏆 Top Performer: {top_performer['Store Name']} (₹{int(top_performer['MTD Value']):,})",
                                    formats['data_normal'])

            # RBM SHEETS
            rbm_headers = ['Store Name', 'MTD Value Conversion', 'FTD Value Conversion', 'MTD Count', 'FTD Count', 'MTD Value', 'FTD Value', 'PREV MONTH SALE', 'DIFF %', 'ASP']
            for rbm in report_df['RBM'].dropna().unique():
                rbm_data = report_df[report_df['RBM'] == rbm].sort_values('MTD Value', ascending=False)
                worksheet_name = rbm[:31] if len(rbm) > 31 else rbm
                rbm_ws = workbook.add_worksheet(worksheet_name)

                # Optimized: Use fixed column widths for speed
                rbm_column_widths = [25, 20, 20, 12, 12, 12, 12, 18, 10, 15]
                for i, width in enumerate(rbm_column_widths):
                    rbm_ws.set_column(i, i, width)

                # RBM Title and subtitle
                rbm_ws.merge_range(0, 0, 0, len(rbm_headers) - 1, f" {rbm} - Sales Performance Report", formats['rbm_title'])
                rbm_ws.merge_range(1, 0, 1, len(rbm_headers) - 1, f"Report Period: {ist_time.strftime('%B %Y')} | Generated: {ist_time.strftime('%d %B %Y %I:%M %p IST')}", formats['rbm_subtitle'])

                # RBM Summary
                rbm_total_stores = len(rbm_data)
                rbm_active_stores = len(rbm_data[rbm_data['FTD Count'] > 0])
                rbm_inactive_stores = rbm_total_stores - rbm_active_stores
                rbm_total_amount = rbm_data['MTD Value'].sum()
                rbm_ws.merge_range(3, 0, 3, 1, "📈 PERFORMANCE OVERVIEW", formats['rbm_summary'])
                rbm_ws.merge_range(3, 2, 3, len(rbm_headers) - 1, f"Total Stores: {rbm_total_stores} | Active: {rbm_active_stores} | Inactive: {rbm_inactive_stores} | Total Revenue: ₹{rbm_total_amount:,}", formats['rbm_summary'])

                # Best performer
                if len(rbm_data) > 0:
                    best_performer = rbm_data.iloc[0]
                    rbm_ws.merge_range(4, 0, 4, len(rbm_headers) - 1, f"🥇 Best Performer: {best_performer['Store Name']} - ₹{int(best_performer['MTD Value']):,}", formats['rbm_performance'])

                # Headers
                for col, header in enumerate(rbm_headers):
                    rbm_ws.write(6, col, header, formats['rbm_header'])

                # Data rows (OPTIMIZED)
                for row_idx, (_, row) in enumerate(rbm_data.iterrows(), start=7):
                    is_alternate = (row_idx - 7) % 2 == 1
                    data_format = formats['rbm_data_alternate'] if is_alternate else formats['rbm_data_normal']
                    store_format = formats['rbm_store_name_alt'] if is_alternate else formats['rbm_store_name']
                    asp_format = formats['asp_format_alt'] if is_alternate else formats['asp_format']
                    
                    # Prepare row data - write entire row at once
                    row_data = [
                        row['Store Name'],
                        row['MTD Value Conversion'] / 100,
                        row['FTD Value Conversion'] / 100,
                        int(row['MTD Count']),
                        int(row['FTD Count']),
                        int(row['MTD Value']),
                        int(row['FTD Value']),
                        int(row['PREV MONTH SALE']),
                        f"{row['DIFF %']}%",
                        row['ASP']
                    ]
                    
                    # Write entire row
                    rbm_ws.write_row(row_idx, 0, row_data, data_format)
                    
                    # Override specific cells
                    rbm_ws.write(row_idx, 0, row['Store Name'], store_format)
                    
                    # MTD Conversion - conditional
                    mtd_conversion = row['MTD Value Conversion']
                    conversion_format = formats['rbm_conversion_format_alt'] if is_alternate else formats['rbm_conversion_format']
                    if mtd_conversion > 2:
                        rbm_ws.write(row_idx, 1, mtd_conversion / 100, formats['rbm_conversion_green'])
                    elif mtd_conversion < 2:
                        rbm_ws.write(row_idx, 1, mtd_conversion / 100, formats['rbm_conversion_low'])
                    else:
                        rbm_ws.write(row_idx, 1, mtd_conversion / 100, conversion_format)

                    # FTD Conversion - conditional
                    ftd_conversion = row['FTD Value Conversion']
                    if ftd_conversion > 2:
                        rbm_ws.write(row_idx, 2, ftd_conversion / 100, formats['rbm_conversion_green'])
                    elif ftd_conversion < 2:
                        rbm_ws.write(row_idx, 2, ftd_conversion / 100, formats['rbm_conversion_low'])
                    else:
                        rbm_ws.write(row_idx, 2, ftd_conversion / 100, conversion_format)
                    
                    # ASP with currency format
                    rbm_ws.write(row_idx, 9, row['ASP'], asp_format)

                # RBM Total row
                total_row = len(rbm_data) + 8
                rbm_ws.write(total_row, 0, '🎯 TOTAL', formats['rbm_total_label'])
                rbm_total_mtd_conversion = round((rbm_data['MTD Value'].sum() / rbm_data['Product_MTD_Amount'].sum()) * 100, 2) if rbm_data['Product_MTD_Amount'].sum() != 0 else 0
                rbm_ws.write(total_row, 1, f"{rbm_total_mtd_conversion}%", formats['rbm_total'])
                rbm_total_ftd_conversion = round((rbm_data['FTD Value'].sum() / rbm_data['Product_FTD_Amount'].sum()) * 100, 2) if rbm_data['Product_FTD_Amount'].sum() != 0 else 0
                rbm_ws.write(total_row, 2, f"{rbm_total_ftd_conversion}%", formats['rbm_total'])
                rbm_ws.write(total_row, 3, rbm_data['MTD Count'].sum(), formats['rbm_total'])
                rbm_ws.write(total_row, 4, rbm_data['FTD Count'].sum(), formats['rbm_total'])
                rbm_ws.write(total_row, 5, rbm_data['MTD Value'].sum(), formats['rbm_total'])
                rbm_ws.write(total_row, 6, rbm_data['FTD Value'].sum(), formats['rbm_total'])
                rbm_ws.write(total_row, 7, rbm_data['PREV MONTH SALE'].sum(), formats['rbm_total'])
                total_prev = rbm_data['PREV MONTH SALE'].sum()
                total_curr = rbm_data['MTD Value'].sum()
                overall_growth = round(((total_curr - total_prev) / total_prev) * 100, 2) if total_prev != 0 else 0
                rbm_ws.write(total_row, 8, f"{overall_growth}%", formats['rbm_total'])
                overall_asp = round(rbm_data['MTD Value'].sum() / rbm_data['MTD Count'].sum(), 2) if rbm_data['MTD Count'].sum() != 0 else 0
                rbm_ws.write(total_row, 9, overall_asp, formats['asp_total'])

                # RBM Insights
                insights_row = total_row + 2
                if overall_growth > 15:
                    rbm_ws.merge_range(insights_row, 0, insights_row, len(rbm_headers) - 1,
                                     f"📈 Excellent Growth: {overall_growth}% increase from previous month",
                                     formats['rbm_summary'])
                elif overall_growth < 0:
                    rbm_ws.merge_range(insights_row, 0, insights_row, len(rbm_headers) - 1,
                                     f"📉 Needs Attention: {abs(overall_growth)}% decrease from previous month",
                                     formats['rbm_summary'])
                else:
                    rbm_ws.merge_range(insights_row, 0, insights_row, len(rbm_headers) - 1,
                                     f"📊 Stable Performance: Less change from previous month",
                                     formats['rbm_summary'])

                # Top 3 performers
                insights_row += 1
                top_3_stores = rbm_data.head(3)
                if len(top_3_stores) > 0:
                    top_stores_text = " | ".join([f"{store['Store Name']}: ₹{int(store['MTD Value']):,}"
                                                for _, store in top_3_stores.iterrows()])
                    rbm_ws.merge_range(insights_row, 0, insights_row, len(rbm_headers) - 1,
                                     f"🏆 Top 3 Performers: {top_stores_text}",
                                     formats['rbm_summary'])

        excel_output.seek(0)
        gc.collect()  # Free memory after heavy Excel processing
        return send_file(
            excel_output,
            download_name=f"OSG_Sales_Report_{today.strftime('%Y%m%d')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True
        )

    except Exception as e:
        print(f"Report 1 Error: {e}")
        import traceback
        traceback.print_exc()
        flash(f"Error generating report: {str(e)}", "error")
        return redirect(url_for('reports_tools'))

@app.route('/reports/generate_2', methods=['POST'])
@login_required
def generate_report_2():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    try:
        selected_date_str = request.form.get('selected_date', '').strip()
        time_slot_raw     = request.form.get('time_slot', '').strip()
        book2_file        = request.files.get('book2')

        # --- INPUT VALIDATION ---
        try:
            selected_date_str = validate_date_str(selected_date_str, field_name='selected_date')
            if not selected_date_str:
                flash("Report date is required.", "error")
                return redirect(url_for('reports_tools'))
            time_slot = validate_time_slot(time_slot_raw)
        except ValueError as e:
            flash(str(e), "error")
            return redirect(url_for('reports_tools'))

        if not book2_file:
            flash("Daily Sales Report file is required.", "error")
            return redirect(url_for('reports_tools'))

        try:
            validate_file_upload(book2_file, field_name='Daily Sales Report')
        except ValueError as e:
            flash(str(e), "error")
            return redirect(url_for('reports_tools'))
        
        # --- END VALIDATION ---

        from services.pg_sync import _get_connection
        import pandas as pd
        
        # Feature Store List
        future_path = os.path.join(BASE_DIR, "Future Store List.xlsx")
        try:
            conn = _get_connection()
            future_df = pd.read_sql('SELECT store as "Store" FROM future_store_list', conn)
            if future_df.empty: raise Exception("Empty DB table future_store_list")
        except Exception as e:
            if not os.path.exists(future_path):
                 flash("Future Store List.xlsx not found on server or DB.", "error")
                 return redirect(url_for('reports_tools'))
            try:
                future_df = timed_excel_read(future_path, "Future Store List", engine='openpyxl')
            except:
                future_df = timed_excel_read(future_path, engine='openpyxl')

        future_df = future_df.loc[:, ~future_df.columns.duplicated()]
        future_df = future_df.loc[:, ~future_df.columns.duplicated()]

        try:
            book2_df = timed_excel_read(book2_file, "Sales Data", engine='openpyxl')
        except:
            book2_file.seek(0)
            book2_df = timed_excel_read(book2_file, engine='openpyxl')

        book2_df = book2_df.loc[:, ~book2_df.columns.duplicated()]

        # Map 'Branch' to 'Store' safely if needed
        if 'Branch' in book2_df.columns and 'Store' not in book2_df.columns:
            book2_df.rename(columns={'Branch': 'Store'}, inplace=True)
        elif 'Store Name' in book2_df.columns and 'Store' not in book2_df.columns:
            book2_df.rename(columns={'Store Name': 'Store'}, inplace=True)

        # Check required columns
        required_cols = ['Store', 'QUANTITY', 'AMOUNT']
        missing_cols = [c for c in required_cols if c not in book2_df.columns]
        if missing_cols:
             raise ValueError(f"Uploaded file is missing required columns: {', '.join(missing_cols)}")
        
        agg = book2_df.groupby('Store', as_index=False).agg({
            'QUANTITY': 'sum',
            'AMOUNT': 'sum'
        })
        
        # Ensure Store column in future_df exists
        if 'Store' not in future_df.columns:
             # try finding branch
             if 'Branch' in future_df.columns:
                 future_df.rename(columns={'Branch': 'Store'}, inplace=True)
             else:
                 raise ValueError("Future Store List missing 'Store' column.")

        all_stores = pd.DataFrame(pd.concat([future_df['Store'], agg['Store']]).unique(), columns=['Store'])
        merged = all_stores.merge(agg, on='Store', how='left')
        merged['QUANTITY'] = merged['QUANTITY'].fillna(0).astype(int)
        merged['AMOUNT'] = merged['AMOUNT'].fillna(0).astype(int)
        
        merged = merged.sort_values(by='AMOUNT', ascending=False).reset_index(drop=True)
        
        total = pd.DataFrame([{
            'Store': 'TOTAL',
            'QUANTITY': merged['QUANTITY'].sum(),
            'AMOUNT': merged['AMOUNT'].sum()
        }])
        
        final_df = pd.concat([merged, total], ignore_index=True)
        final_df.rename(columns={'Store': 'Branch'}, inplace=True)

        # Excel Generation with OpenPyXL
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Store Report"
        
        title_text = f"{selected_date_str} EW Sale Till {time_slot}"
        ws.merge_cells('A1:C1')
        ws['A1'] = title_text
        ws['A1'].font = Font(bold=True, size=11, color="FFFFFF")
        ws['A1'].fill = PatternFill("solid", fgColor="4F81BD")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        header_fill = PatternFill("solid", fgColor="4F81BD")
        data_fill = PatternFill("solid", fgColor="DCE6F1")
        red_fill = PatternFill("solid", fgColor="F4CCCC")
        total_fill = PatternFill("solid", fgColor="10B981")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=True), start=2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 2: # Header
                    cell.fill = header_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                elif row[0] == 'TOTAL':
                    cell.fill = total_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                elif (c_idx == 1 and row[1] <= 0) or ((c_idx == 2 or c_idx == 3) and value <= 0): # Branch, Qty, Amt check
                    cell.fill = red_fill
                else:
                    cell.fill = data_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
        
        # approximate cleanup
        wb.save(buffer)
        buffer.seek(0)
        gc.collect()  # Free memory after Excel processing
        return send_file(
            buffer,
            download_name=f"Store_Summary_{selected_date_str}_{time_slot}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True
        )

    except Exception as e:
        print(f"Report 2 Error: {e}")
        flash(f"Error generating report: {str(e)}", "error")
        return redirect(url_for('reports_tools'))

@app.route('/reports/map_data', methods=['POST'])
@login_required
def map_data():
    try:
        osg_file     = request.files.get('osg_file')
        product_file = request.files.get('product_file')

        if not osg_file or not product_file:
            flash("Both OSG and Product files are required.", "error")
            return redirect(url_for('reports_tools'))

        # --- INPUT VALIDATION ---
        for upload, label in [(osg_file, 'OSG file'), (product_file, 'Product file')]:
            try:
                validate_file_upload(upload, field_name=label)
            except ValueError as e:
                flash(str(e), "error")
                return redirect(url_for('reports_tools'))
        # --- END VALIDATION ---

        osg_df     = pd.read_excel(osg_file)
        product_df = pd.read_excel(product_file)
        
        # SKU Mapping
        sku_category_mapping = {
            "Warranty : Water Cooler/Dispencer/Geyser/RoomCooler/Heater": [
                "COOLER", "DISPENCER", "GEYSER", "ROOM COOLER", "HEATER", "WATER HEATER", "WATER DISPENSER"
            ],
            "Warranty : Fan/Mixr/IrnBox/Kettle/OTG/Grmr/Geysr/Steamr/Inductn": [
                "FAN", "MIXER", "IRON BOX", "KETTLE", "OTG", "GROOMING KIT", "GEYSER", "STEAMER", "INDUCTION",
                "CEILING FAN", "TOWER FAN", "PEDESTAL FAN", "INDUCTION COOKER", "ELECTRIC KETTLE", "WALL FAN", "MIXER GRINDER", "CELLING FAN"
            ],
            "AC : EWP : Warranty : AC": ["AC", "AIR CONDITIONER", "AC INDOOR"],
            "HAEW : Warranty : Air Purifier/WaterPurifier": ["AIR PURIFIER", "WATER PURIFIER"],
            "HAEW : Warranty : Dryer/MW/DishW": ["DRYER", "MICROWAVE OVEN", "DISH WASHER", "MICROWAVE OVEN-CONV"],
            "HAEW : Warranty : Ref/WM": [
                "REFRIGERATOR", "WASHING MACHINE", "WASHING MACHINE-TL", "REFRIGERATOR-DC",
                "WASHING MACHINE-FL", "WASHING MACHINE-SA", "REF", "REFRIGERATOR-CBU", "REFRIGERATOR-FF", "WM"
            ],
            "HAEW : Warranty : TV": ["TV", "TV 28 %", "TV 18 %"],
            "TV : TTC : Warranty and Protection : TV": ["TV", "TV 28 %", "TV 18 %"],
            "TV : Spill and Drop Protection": ["TV", "TV 28 %", "TV 18 %"],
            "HAEW : Warranty :Chop/Blend/Toast/Air Fryer/Food Processr/JMG/Induction": [
                "CHOPPER", "BLENDER", "TOASTER", "AIR FRYER", "FOOD PROCESSOR", "JUICER", "INDUCTION COOKER"
            ],
            "HAEW : Warranty : HOB and Chimney": ["HOB", "CHIMNEY"],
            "HAEW : Warranty : HT/SoundBar/AudioSystems/PortableSpkr": [
                "HOME THEATRE", "AUDIO SYSTEM", "SPEAKER", "SOUND BAR", "PARTY SPEAKER"
            ],
            "HAEW : Warranty : Vacuum Cleaner/Fans/Groom&HairCare/Massager/Iron": [
                "VACUUM CLEANER", "FAN", "MASSAGER", "IRON BOX", "CEILING FAN", "TOWER FAN", "PEDESTAL FAN", "WALL FAN", "ROBO VACCUM CLEANER"
            ],
            "AC AMC": ["AC", "AC INDOOR"]
        }

        product_df['Category'] = product_df['Category'].str.upper().fillna('')
        product_df['Model'] = product_df['Model'].fillna('')
        product_df['Customer Mobile'] = product_df['Customer Mobile'].astype(str)
        product_df['Invoice Number'] = product_df['Invoice Number'].astype(str)
        product_df['Item Rate'] = pd.to_numeric(product_df['Item Rate'], errors='coerce')
        product_df['IMEI'] = product_df['IMEI'].astype(str).fillna('')
        product_df['Brand'] = product_df['Brand'].fillna('')
        osg_df['Customer Mobile'] = osg_df['Customer Mobile'].astype(str)

        def extract_price_slab(text):
            match = re.search(r"Slab\s*:\s*(\d+)K-(\d+)K", str(text))
            if match:
                return int(match.group(1)) * 1000, int(match.group(2)) * 1000
            return None, None

        def get_model(row):
            mobile = row['Customer Mobile']
            retailer_sku = str(row.get('Retailer SKU', ''))
            invoice = str(row.get('Invoice Number', ''))
            user_products = product_df[product_df['Customer Mobile'] == mobile]

            if user_products.empty:
                return ''
            unique_models = user_products['Model'].dropna().unique()
            if len(unique_models) == 1:
                return unique_models[0]

            mapped_keywords = []
            for sku_key, keywords in sku_category_mapping.items():
                if sku_key in retailer_sku:
                    mapped_keywords = [kw.lower() for kw in keywords]
                    break   

            filtered = user_products[user_products['Category'].str.lower().isin(mapped_keywords)]
            if filtered['Model'].nunique() == 1:
                return filtered['Model'].iloc[0]

            slab_min, slab_max = extract_price_slab(retailer_sku)
            if slab_min and slab_max:
                slab_filtered = filtered[(filtered['Item Rate'] >= slab_min) & (filtered['Item Rate'] <= slab_max)]
                if slab_filtered['Model'].nunique() == 1:
                    return slab_filtered['Model'].iloc[0]
                invoice_filtered = slab_filtered[slab_filtered['Invoice Number'].astype(str) == invoice]
                if invoice_filtered['Model'].nunique() == 1:
                    return invoice_filtered['Model'].iloc[0]

            return ''

        osg_df['Model'] = osg_df.apply(get_model, axis=1)
        category_brand_df = product_df[['Customer Mobile', 'Model', 'Category', 'Brand']].drop_duplicates()
        osg_df = osg_df.merge(category_brand_df, on=['Customer Mobile', 'Model'], how='left')

        invoice_pool = defaultdict(list)
        itemrate_pool = defaultdict(list)
        imei_pool = defaultdict(list)

        for _, row in product_df.iterrows():
            key = (row['Customer Mobile'], row['Model'])
            invoice_pool[key].append(row['Invoice Number'])
            itemrate_pool[key].append(row['Item Rate'])
            imei_pool[key].append(row['IMEI'])

        invoice_usage_counter = defaultdict(int)
        itemrate_usage_counter = defaultdict(int)
        imei_usage_counter = defaultdict(int)

        def assign_from_pool(row, pool, counter_dict):
            key = (row['Customer Mobile'], row['Model'])
            values = pool.get(key, [])
            index = counter_dict[key]
            if index < len(values):
                counter_dict[key] += 1
                return values[index]
            return ''

        osg_df['Product Invoice Number'] = osg_df.apply(lambda row: assign_from_pool(row, invoice_pool, invoice_usage_counter), axis=1)
        osg_df['Item Rate'] = osg_df.apply(lambda row: assign_from_pool(row, itemrate_pool, itemrate_usage_counter), axis=1)
        osg_df['IMEI'] = osg_df.apply(lambda row: assign_from_pool(row, imei_pool, imei_usage_counter), axis=1)
        osg_df['Store Code'] = osg_df['Product Invoice Number'].astype(str).apply(
            lambda x: re.search(r'\b([A-Z]{2,})\b', x).group(1) if re.search(r'\b([A-Z]{2,})\b', x) else ''
        )

        def extract_warranty_duration(sku):
            sku = str(sku)
            match = re.search(r'Dur\s*:\s*(\d+)\+(\d+)', sku)
            if match:
                return int(match.group(1)), int(match.group(2))
            match = re.search(r'(\d+)\+(\d+)\s*SDP-(\d+)', sku)
            if match:
                return int(match.group(1)), f"{match.group(3)}P+{match.group(2)}W"
            match = re.search(r'Dur\s*:\s*(\d+)', sku)
            if match:
                return 1, int(match.group(1))
            match = re.search(r'(\d+)\+(\d+)', sku)
            if match:
                return int(match.group(1)), int(match.group(2))
            return '', ''

        osg_df[['Manufacturer Warranty', 'Duration (Year)']] = osg_df['Retailer SKU'].apply(
            lambda sku: pd.Series(extract_warranty_duration(sku))
        )

        def highlight_row(row):
            missing_fields = pd.isna(row.get('Model')) or str(row.get('Model')).strip() == ''
            missing_fields |= pd.isna(row.get('IMEI')) or str(row.get('IMEI')).strip() == ''
            try:
                if float(row.get('Plan Price', 0)) < 0:
                    missing_fields |= True
            except:
                missing_fields |= True
            return ['background-color: lightblue'] * len(row) if missing_fields else [''] * len(row)

        final_columns = [
            'Customer Mobile', 'Date', 'Invoice Number','Product Invoice Number', 'Customer Name', 'Store Code', 'Branch', 'Region',
            'IMEI', 'Category', 'Brand', 'Quantity', 'Item Code', 'Model', 'Plan Type', 'EWS QTY', 'Item Rate',
            'Plan Price', 'Sold Price', 'Email', 'Product Count', 'Manufacturer Warranty', 'Retailer SKU', 'OnsiteGo SKU',
            'Duration (Year)', 'Total Coverage', 'Comment', 'Return Flag', 'Return against invoice No.',
            'Primary Invoice No.'
        ]

        for col in final_columns:
            if col not in osg_df.columns:
                osg_df[col] = ''
        osg_df['Quantity'] = 1
        osg_df['EWS QTY'] = 1
        osg_df = osg_df[final_columns]

        def convert_df(df):
           output = io.BytesIO()
           styled_df = df.style.apply(highlight_row, axis=1)
           with pd.ExcelWriter(output, engine='openpyxl') as writer:
            styled_df.to_excel(writer, index=False)
           output.seek(0)
           return output

        excel_data = convert_df(osg_df)
        
        flash("✅ Data Mapping Completed Successfully. The OSG and product data has been successfully mapped. The report download will begin shortly.", "success")

        return send_file(
            excel_data,
            download_name="OSG_Product_Mapping_Report.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True
        )
    except Exception as e:
        print(f"Mapping Error: {e}")
        import traceback
        traceback.print_exc()
        flash(f"Mapping Failed: {str(e)}", "error")
        return redirect(url_for('reports_tools'))

@app.route('/api/export-claims-excel', methods=['POST'])
@login_required
def export_claims_excel():
    """Export filtered claims as Excel (.xlsx) file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        data = request.get_json()

        # --- INPUT VALIDATION ---
        try:
            raw_ids   = data.get('claim_ids', []) if data else []
            claim_ids = validate_claim_ids_list(raw_ids)
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        # --- END VALIDATION ---

        # Fetch all claims from cache
        all_claims = fetch_claims_from_db()

        # Filter to only validated IDs
        if claim_ids:
            id_set = set(claim_ids)
            claims = [c for c in all_claims if str(c.claim_id) in id_set]
        else:
            claims = all_claims

        # --- Build workbook ---
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Claims Export"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(border_style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

        # Header row
        headers = [
            "SR No", "Claim ID", "Submitted Date", "Customer Name", "Mobile",
            "Branch", "Product", "Issue", "Status",
            "Replacement Progress %", "Complete", "Aging Days"
        ]
        ws.append(headers)

        # Style header
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        ws.row_dimensions[1].height = 22

        # Data rows
        def calc_progress(c):
            stages = [
                c.data.get("Customer Confirmation", ""),
                c.data.get("Approval Mail Received From Onsitego (Yes/No)", ""),
                c.data.get("Mail Sent To Store (Yes/No)", ""),
                c.data.get("Invoice Generated (Yes/No)", ""),
                c.data.get("Invoice Sent To Onsitego (Yes/No)", ""),
                c.data.get("Settled With Accounts (Yes/No)", ""),
            ]
            done = sum(1 for s in stages if str(s).lower() == "yes")
            return round((done / len(stages)) * 100)

        # Pending fill for aging
        pending_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        overdue_fill = PatternFill(start_color="FFDCDC", end_color="FFDCDC", fill_type="solid")

        today = datetime.date.today()

        for i, claim in enumerate(claims, 2):
            # Aging Days: only for pending (not complete) claims
            if not claim.complete:
                try:
                    submitted_date = claim.created_at.date() if claim.created_at else None
                    aging_days = (today - submitted_date).days if submitted_date else '-'
                except:
                    aging_days = '-'
            else:
                aging_days = '-'

            row_data = [
                str(claim.sr_no or ''),
                str(claim.claim_id or ''),
                str(claim.created_at.strftime('%d %b %Y') if claim.created_at else ''),
                str(claim.customer_name or ''),
                str(claim.mobile_no or ''),
                str(claim.branch or '-'),
                str(claim.model or ''),
                str(claim.issue or ''),
                str(claim.status or ''),
                calc_progress(claim),
                'Yes' if claim.complete else 'No',
                aging_days
            ]
            ws.append(row_data)

            # Alternate row shading
            fill = alt_fill if i % 2 == 0 else None
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=i, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if fill:
                    cell.fill = fill

            # Color-code Aging Days cell (last column) for pending cases
            if not claim.complete and isinstance(aging_days, int):
                aging_cell = ws.cell(row=i, column=len(headers))
                aging_cell.font = Font(bold=True)
                if aging_days > 7:
                    aging_cell.fill = overdue_fill
                elif aging_days > 3:
                    aging_cell.fill = pending_fill

        # Auto-size columns
        col_widths = [14, 16, 16, 22, 14, 18, 30, 40, 22, 22, 10, 14]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"OSG_Claims_Export_{datetime.date.today().strftime('%Y-%m-%d')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Excel Export Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ----------------------
# POSTGRESQL SYNC ROUTES
# ----------------------

@app.route('/api/pg-status', methods=['GET'])
@admin_required
def api_pg_status():
    """
    Check PostgreSQL connectivity and return row count.
    GET /api/pg-status
    """
    if not PG_AVAILABLE:
        return jsonify({"configured": False, "error": "DATABASE_URL is not set."}), 503

    info = pg_test_connection()
    if info.get("success"):
        try:
            import psycopg2
            conn = psycopg2.connect(os.environ.get("DATABASE_URL"))
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM claims;")
                count = cur.fetchone()[0]
                cur.execute("SELECT MAX(synced_at) FROM claims;")
                last_sync = cur.fetchone()[0]
            conn.close()
            info["row_count"] = count
            info["last_synced_at"] = str(last_sync) if last_sync else "Never"
        except Exception as e:
            info["row_count"] = "unknown"
            info["last_synced_at"] = "unknown"
            info["note"] = str(e)
    return jsonify(info)

@app.route('/api/pg-claims', methods=['GET'])
@admin_required
def api_pg_claims():
    """
    Return all claims stored in PostgreSQL as JSON.
    Useful for debugging. Supports ?limit=N query param.
    GET /api/pg-claims?limit=50
    """
    if not PG_AVAILABLE:
        return jsonify({"success": False, "error": "DATABASE_URL is not configured."}), 503

    try:
        limit = int(request.args.get("limit", 100))
        limit = max(1, min(limit, 1000))  # clamp 1–1000
    except (ValueError, TypeError):
        limit = 100

    try:
        import psycopg2
        import psycopg2.extras
        conn = psycopg2.connect(os.environ.get("DATABASE_URL"))
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM claims ORDER BY synced_at DESC LIMIT %s;", (limit,))
            rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return jsonify({"success": True, "count": len(rows), "claims": rows})
    except Exception as e:
        logging.error(f"[PG_CLAIMS_ROUTE] Error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500



# ── API: bulk import Excel → Dynamic Table ───────────────────────────
@app.route('/api/db/bulk-import/<table_name>', methods=['POST'])
@login_required
def api_db_bulk_import(table_name):
    if not session.get('db_mgr_auth'):
        return jsonify({'error': 'Unauthorized'}), 403
    if table_name not in _DB_ALLOWED_TABLES:
        return jsonify({'error': 'Invalid table'}), 400
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'No file selected'}), 400
    import re, pandas as pd, psycopg2, psycopg2.extras, io

    def pg_col(name):
        s = re.sub(r'[^a-z0-9]+', '_', str(name).strip().lower()).strip('_')
        return s or 'col'

    try:
        df = pd.read_excel(io.BytesIO(f.read()), dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        col_map = {}
        for c in df.columns:
            pg = pg_col(c)
            col_map[c] = pg if pg not in col_map.values() else pg + '_2'
        pg_cols = list(col_map.values())

        conn = psycopg2.connect(os.environ['DATABASE_URL'])
        cur  = conn.cursor()

        # Ensure columns exist
        cur.execute(f"""SELECT column_name FROM information_schema.columns
                       WHERE table_name='{table_name}' AND table_schema='public'""")
        existing_cols = {r[0] for r in cur.fetchall()}
        for pg in pg_cols:
            if pg not in existing_cols and pg not in ('id','imported_at'):
                cur.execute(f'ALTER TABLE {table_name} ADD COLUMN IF NOT EXISTS "{pg}" TEXT')

        # Insert rows
        col_ids  = ', '.join(f'"{c}"' for c in pg_cols)
        placeholders = ', '.join(['%s'] * len(pg_cols))
        sql = f'INSERT INTO {table_name} ({col_ids}) VALUES ({placeholders})'
        rows = []
        for _, row in df.iterrows():
            vals = []
            
            # Find exact keys from df
            remarks_key = next((k for k in col_map.keys() if str(k).strip().lower() == "remarks"), None)
            onsitego_key = next((k for k in col_map.keys() if "onsitego" in str(k).lower() and "status" in str(k).lower()), None)
            notes_key = next((k for k in col_map.keys() if "follow up" in str(k).lower() and "notes" in str(k).lower()), None)
            
            remarks_val = str(row.get(remarks_key, '') if remarks_key else '').strip()
            onsitego_val = str(row.get(onsitego_key, '') if onsitego_key else '').strip()
            notes_val = str(row.get(notes_key, '') if notes_key else '').strip()
            
            if remarks_val.lower() in ('nan', 'none', 'nat'): remarks_val = ""
            if onsitego_val.lower() in ('nan', 'none', 'nat'): onsitego_val = ""
            if notes_val.lower() in ('nan', 'none', 'nat'): notes_val = ""
            
            import datetime
            ts = datetime.datetime.now().strftime('%d/%m/%Y, %I:%M:%S %p').lower()
            
            if remarks_val and remarks_val.lower() not in notes_val.lower():
                notes_val += f"\n[{ts}] [REMARK]: {remarks_val}"
            if onsitego_val and onsitego_val.lower() not in notes_val.lower():
                notes_val += f"\n[{ts}] [ONSITEGO STATUS]: {onsitego_val}"
            
            notes_val = notes_val.strip()
            
            for orig in col_map.keys():
                if orig == notes_key:
                    v = notes_val
                else:
                    v = str(row.get(orig, '') or '').strip()
                vals.append(None if str(v).lower() in ('nan','none','nat','') else v)
            rows.append(tuple(vals))

        psycopg2.extras.execute_batch(cur, sql, rows, page_size=500)
        conn.commit()
        inserted = len(rows)

        cur.execute(f'SELECT COUNT(*) FROM {table_name}')
        total = cur.fetchone()[0]
        conn.close()

        return jsonify({'success': True, 'inserted': inserted, 'total': total, 'filename': f.filename})
    except Exception as e:
        logging.error(f'[BULK_IMPORT] {e}')
        return jsonify({'error': str(e)}), 500

# ─────────────────────────────────────────────────────────────────
# DATABASE MANAGER  — password-protected CRUD for all PG tables
# ─────────────────────────────────────────────────────────────────
_DB_ALLOWED_TABLES = {'claims', 'osid_data', 'myg_all_store', 'future_store_list', 'rbm_bdm_branch'}
_DB_MGR_PASSWORD   = os.environ.get('DB_MANAGER_PASSWORD', 'DBAdmin@2026')

def _dbmgr_conn():
    import psycopg2, psycopg2.extras
    return psycopg2.connect(os.environ['DATABASE_URL'])

@app.route('/db-manager')
@login_required
def db_manager():
    if session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    if not session.get('db_mgr_auth'):
        return redirect(url_for('db_manager_login'))
    return render_template('db_manager.html')

@app.route('/db-manager/login', methods=['GET', 'POST'])
@login_required
def db_manager_login():
    if session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    error = None
    if request.method == 'POST':
        pwd = request.form.get('password', '')
        if pwd == _DB_MGR_PASSWORD:
            session['db_mgr_auth'] = True
            return redirect(url_for('db_manager'))
        error = 'Incorrect password.'
    return render_template('db_manager_login.html', error=error)

@app.route('/db-manager/logout')
def db_manager_logout():
    session.pop('db_mgr_auth', None)
    return redirect(url_for('db_manager_login'))

# ── API: list tables ──────────────────────────────────────────────
@app.route('/api/db/tables')
@login_required
def api_db_tables():
    if not session.get('db_mgr_auth'):
        return jsonify({'error': 'Unauthorized'}), 403
    import psycopg2
    try:
        conn = _dbmgr_conn()
        cur = conn.cursor()
        results = []
        for tbl in sorted(_DB_ALLOWED_TABLES):
            cur.execute(f"SELECT COUNT(*) FROM {tbl}")
            count = cur.fetchone()[0]
            cur.execute("""
                SELECT column_name, data_type
                FROM information_schema.columns
                WHERE table_name = %s AND table_schema = 'public'
                ORDER BY ordinal_position
            """, (tbl,))
            cols = [{'name': r[0], 'type': r[1]} for r in cur.fetchall()]
            results.append({'table': tbl, 'count': count, 'columns': cols})
        conn.close()
        return jsonify({'tables': results})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── API: get rows (paginated) ─────────────────────────────────────
@app.route('/api/db/table/<table_name>')
@login_required
def api_db_get_rows(table_name):
    if not session.get('db_mgr_auth'):
        return jsonify({'error': 'Unauthorized'}), 403
    if table_name not in _DB_ALLOWED_TABLES:
        return jsonify({'error': 'Table not allowed'}), 400
    import psycopg2, psycopg2.extras
    page     = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 100))
    search   = request.args.get('search', '').strip()
    offset   = (page - 1) * per_page
    try:
        conn = _dbmgr_conn()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        # columns
        cur.execute("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name=%s AND table_schema='public'
            ORDER BY ordinal_position
        """, (table_name,))
        cols = [r['column_name'] for r in cur.fetchall()]
        # total count
        if search:
            where = " OR ".join([f"CAST(\"{c}\" AS TEXT) ILIKE %s" for c in cols])
            params = [f'%{search}%'] * len(cols)
            cur.execute(f"SELECT COUNT(*) FROM {table_name} WHERE {where}", params)
            total = cur.fetchone()['count']
            cur.execute(f"SELECT * FROM {table_name} WHERE {where} LIMIT %s OFFSET %s",
                        params + [per_page, offset])
        else:
            cur.execute(f"SELECT COUNT(*) FROM {table_name}")
            total = cur.fetchone()['count']
            cur.execute(f"SELECT * FROM {table_name} LIMIT %s OFFSET %s", (per_page, offset))
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return jsonify({'columns': cols, 'rows': rows, 'total': total, 'page': page, 'per_page': per_page})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── API: update a row ─────────────────────────────────────────────
@app.route('/api/db/table/<table_name>/row/<row_id>', methods=['PUT'])
@login_required
def api_db_update_row(table_name, row_id):
    if not session.get('db_mgr_auth'):
        return jsonify({'error': 'Unauthorized'}), 403
    if table_name not in _DB_ALLOWED_TABLES:
        return jsonify({'error': 'Table not allowed'}), 400
    pk = 'claim_id' if table_name == 'claims' else 'id'
    data = request.json or {}
    data.pop(pk, None)   # don't update primary key
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    set_clause = ", ".join([f'"{k}" = %s' for k in data.keys()])
    vals = list(data.values()) + [row_id]
    try:
        conn = _dbmgr_conn()
        cur = conn.cursor()
        cur.execute(f'UPDATE {table_name} SET {set_clause} WHERE "{pk}" = %s', vals)
        conn.commit(); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── API: delete a row ─────────────────────────────────────────────
@app.route('/api/db/table/<table_name>/row/<row_id>', methods=['DELETE'])
@login_required
def api_db_delete_row(table_name, row_id):
    if not session.get('db_mgr_auth'):
        return jsonify({'error': 'Unauthorized'}), 403
    if table_name not in _DB_ALLOWED_TABLES:
        return jsonify({'error': 'Table not allowed'}), 400
    pk = 'claim_id' if table_name == 'claims' else 'id'
    try:
        conn = _dbmgr_conn()
        cur = conn.cursor()
        cur.execute(f'DELETE FROM {table_name} WHERE "{pk}" = %s', (row_id,))
        conn.commit(); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── API: add a row ────────────────────────────────────────────────
@app.route('/api/db/table/<table_name>/row', methods=['POST'])
@login_required
def api_db_add_row(table_name):
    if not session.get('db_mgr_auth'):
        return jsonify({'error': 'Unauthorized'}), 403
    if table_name not in _DB_ALLOWED_TABLES:
        return jsonify({'error': 'Table not allowed'}), 400
    data = request.json or {}
    data.pop('id', None); data.pop('imported_at', None)
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    cols_sql = ", ".join([f'"{k}"' for k in data.keys()])
    vals_sql = ", ".join(["%s"] * len(data))
    try:
        conn = _dbmgr_conn()
        cur = conn.cursor()
        cur.execute(f'INSERT INTO {table_name} ({cols_sql}) VALUES ({vals_sql})', list(data.values()))
        conn.commit(); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── API: delete all rows ──────────────────────────────────────────
@app.route('/api/db/table/<table_name>/all', methods=['DELETE'])
@login_required
def api_db_delete_all(table_name):
    if not session.get('db_mgr_auth'): return jsonify({'error': 'Unauthorized'}), 403
    if table_name not in _DB_ALLOWED_TABLES: return jsonify({'error': 'Table not allowed'}), 400
    
    data = request.json or {}
    if data.get('password') != _DB_MGR_PASSWORD:
        return jsonify({'error': 'Invalid DB Manager password'}), 401

    try:
        conn = _dbmgr_conn()
        cur = conn.cursor()
        cur.execute(f'TRUNCATE TABLE {table_name} RESTART IDENTITY CASCADE')
        conn.commit(); conn.close()
        logging.warning(f"[DBM] All data truncated from {table_name} by user: {session.get('user')}")
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
# ── BACKGROUND SHEET POLLER ───────────────────────────────────────
def start_sheet_poller():
    def poller():
        import time
        import requests
        import threading
        
        while True:
            time.sleep(15)  # Poll every 15 seconds
            try:
                web_app_url = os.environ.get("WEB_APP_URL")
                if not web_app_url or not PG_AVAILABLE:
                    continue
                
                resp = requests.get(web_app_url, timeout=30)
                if resp.status_code != 200:
                    continue
                sheet_data = resp.json()
                
                # Fetch current db claims
                from services.pg_sync import fetch_claims_from_postgres, upsert_claim_to_postgres
                db_data = fetch_claims_from_postgres()
                db_dict = {str(c.get("Claim ID", "")): c for c in db_data}
                
                updates_made = False
                for row in sheet_data:
                    cid = str(row.get("Claim ID") or row.get("claim_id") or "").strip()
                    # Fix space vs hyphen mismatch if user manually edited Claim ID
                    cid = cid.replace(" ", "-") 
                    if not cid or not cid.startswith("CLM"):
                        continue
                        
                    db_row = db_dict.get(cid, {})
                    
                    # Extract sheet Remarks, Onsitego, and SR No
                    s_remarks = ""
                    s_onsitego = ""
                    s_sr_no = ""
                    for k, v in row.items():
                        if str(k).strip().lower() == "remarks":
                            s_remarks = str(v).strip()
                        elif "onsitego" in str(k).lower() and "status" in str(k).lower():
                            s_onsitego = str(v).strip()
                        elif str(k).strip().lower() == "sr no" or str(k).strip().lower() == "sr_no":
                            s_sr_no = str(v).strip()
                            
                    # Extract DB Remarks, Onsitego, and SR No
                    db_remarks = ""
                    db_onsitego = ""
                    db_sr_no = ""
                    for k, v in db_row.items():
                        if str(k).strip().lower() == "remarks":
                            db_remarks = str(v).strip()
                        elif "onsitego" in str(k).lower() and "status" in str(k).lower():
                            db_onsitego = str(v).strip()
                        elif str(k).strip().lower() == "sr no" or str(k).strip().lower() == "sr_no":
                            db_sr_no = str(v).strip()
                            
                    if s_remarks.lower() in ('nan', 'none', 'nat'): s_remarks = ""
                    if s_onsitego.lower() in ('nan', 'none', 'nat'): s_onsitego = ""
                    if s_sr_no.lower() in ('nan', 'none', 'nat'): s_sr_no = ""
                    if db_remarks.lower() in ('nan', 'none', 'nat'): db_remarks = ""
                    if db_onsitego.lower() in ('nan', 'none', 'nat'): db_onsitego = ""
                    if db_sr_no.lower() in ('nan', 'none', 'nat'): db_sr_no = ""

                    # If changed in sheet, push to DB!
                    if (s_remarks and s_remarks.lower() != db_remarks.lower()) or \
                       (s_onsitego and s_onsitego.lower() != db_onsitego.lower()) or \
                       (s_sr_no and s_sr_no.lower() != db_sr_no.lower()):
                        print(f"[POLLER] Detected change in sheet for Claim {cid} - Syncing partial data to DB")
                        
                        # Build a partial dictionary with only the fields we want to pull from Google Sheets
                        partial_row = {
                            "Claim ID": cid
                        }
                        if s_remarks and s_remarks.lower() != db_remarks.lower():
                            partial_row["Remarks"] = s_remarks
                        if s_onsitego and s_onsitego.lower() != db_onsitego.lower():
                            partial_row["ONSITEGO - STATUS"] = s_onsitego
                        if s_sr_no and s_sr_no.lower() != db_sr_no.lower():
                            partial_row["SR No"] = s_sr_no
                        
                        upsert_claim_to_postgres(partial_row)
                        updates_made = True
                        
                if updates_made:
                    global CLAIMS_CACHE
                    CLAIMS_CACHE['last_updated'] = 0
            except Exception as e:
                pass
                
    import threading
    t = threading.Thread(target=poller, daemon=True)
    t.start()

start_sheet_poller()


# ==========================================
# WHATSAPP REPORTS (TELFINY API)
# ==========================================

import csv
import io

@app.route('/whatsapp-reports')
@login_required
def whatsapp_reports():
    return render_template('whatsapp_reports.html')

@app.route('/api/webhooks/telfiny', methods=['POST'])
def telfiny_webhook():
    data = request.json
    if not data:
        return jsonify({'success': False}), 400
    
    msg_id = data.get('messageId')
    mobile = data.get('mobileNumber') or data.get('mobile')
    status = data.get('status', '').lower()
    reason = data.get('errorDescription') or data.get('reason') or ''
    
    if msg_id and mobile:
        try:
            conn = psycopg2.connect(os.environ.get('DATABASE_URL'))
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO whatsapp_message_logs (message_id, mobile_number, status, failure_reason)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (message_id) 
                    DO UPDATE SET status = EXCLUDED.status, failure_reason = EXCLUDED.failure_reason, updated_at = CURRENT_TIMESTAMP
                """, (msg_id, mobile, status, reason))
            conn.commit()
            conn.close()
        except Exception as e:
            logging.error(f'Webhook DB Error: {e}')
            
    return jsonify({'success': True}), 200

@app.route('/api/whatsapp/request-report', methods=['POST'])
@login_required
def request_whatsapp_report():
    data = request.json
    from_date = data.get('fromDate')
    to_date = data.get('toDate')

    if not from_date or not to_date:
        return jsonify({'success': False, 'message': 'Missing dates'}), 400

    try:
        conn = psycopg2.connect(os.environ.get('DATABASE_URL'))
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT status, COUNT(*) as count 
                FROM whatsapp_message_logs 
                WHERE created_at::date >= %s AND created_at::date <= %s
                GROUP BY status
            """, (from_date, to_date))
            
            stats_rows = cur.fetchall()
            stats = {row['status']: row['count'] for row in stats_rows}
            
            cur.execute("""
                SELECT mobile_number as mobile, TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') as date, failure_reason as reason 
                FROM whatsapp_message_logs 
                WHERE created_at::date >= %s AND created_at::date <= %s AND status IN ('failed', 'error')
                ORDER BY created_at DESC
                LIMIT 100
            """, (from_date, to_date))
            failed_messages = cur.fetchall()
            
        conn.close()
        
        return jsonify({
            'success': True,
            'fileID': f'local_{from_date}_{to_date}',
            'status': 'completed',
            'data': {
                'sent': sum(stats.values()), # Total of all statuses = total messages triggered
                'delivered': stats.get('delivered', 0) + stats.get('read', 0),
                'read': stats.get('read', 0),
                'failed': stats.get('failed', 0) + stats.get('error', 0),
                'failed_messages': failed_messages
            }
        }), 200
    except Exception as e:
        logging.error(f'Error querying DB for report: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/whatsapp/poll-report/<file_id>', methods=['GET'])
@login_required
def poll_whatsapp_report(file_id):
    if file_id.startswith('local_'):
        parts = file_id.split('_')
        from_date = parts[1]
        to_date = parts[2]
        
        try:
            conn = psycopg2.connect(os.environ.get('DATABASE_URL'))
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("""
                    SELECT status, COUNT(*) as count 
                    FROM whatsapp_message_logs 
                    WHERE created_at::date >= %s AND created_at::date <= %s
                    GROUP BY status
                """, (from_date, to_date))
                
                stats_rows = cur.fetchall()
                stats = {row['status']: row['count'] for row in stats_rows}
                
                cur.execute("""
                    SELECT mobile_number as mobile, TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') as date, failure_reason as reason 
                    FROM whatsapp_message_logs 
                    WHERE created_at::date >= %s AND created_at::date <= %s AND status IN ('failed', 'error')
                    ORDER BY created_at DESC
                    LIMIT 100
                """, (from_date, to_date))
                failed_messages = cur.fetchall()
                
            conn.close()
            
            return jsonify({
                'status': 'completed',
                'data': {
                    'sent': stats.get('sent', 0) + stats.get('submitted', 0),
                    'delivered': stats.get('delivered', 0),
                    'read': stats.get('read', 0),
                    'failed': stats.get('failed', 0) + stats.get('error', 0),
                    'failed_messages': failed_messages
                }
            }), 200
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500
    return jsonify({'success': False, 'message': 'Invalid file ID'}), 400

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 3000))
    app.run(
        host='0.0.0.0',
        port=port,
        debug=os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'  # Never True in production
    )

